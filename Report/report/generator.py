"""从版本用例 xlsx 生成内网/外网测试总结 HTML 报告。"""

from __future__ import annotations

import base64
import functools
import html
import os
import platform
import re
import subprocess
import sys
import webbrowser
import zlib
from pathlib import Path


def _is_tech_opt_need(title: str) -> bool:
    t = title.strip().rstrip("：:").strip()
    return t == "技术优化需求"


_SUBITEM_PREFIX_RE = re.compile(r"^\s*(?:\d+\s*[)）\.、]|[-*•]\s+)\s*")
# 首表 I2：多行「缺陷链接 / 版本用例链接 / 回归用例链接」+ 未完成&已完成「阻碍～轻微」行（`format_sheet1_i2_text` 生成同结构）
_E2_DEFECT_URL_RE = re.compile(r"缺陷链接\s*[:：]\s*(https?://\S+)")
_E2_VERSION_CASE_URL_RE = re.compile(r"版本用例链接\s*[:：]\s*(https?://\S+)")
_E2_REGRESSION_URL_RE = re.compile(r"回归用例链接\s*[:：]\s*(https?://\S+)")
# 首表 I2：「未完成缺陷」「已完成缺陷」后各级数量（单行内用全角分号 `；` 连接，与 Excel 中格式一致）
_E2_DEFECT_LEVEL_LINE_RE = re.compile(r"(阻碍|严重|一般|轻微)\s*[:：]\s*(\d+)")
_E2_UNFINISHED_BLOCK_RE = re.compile(r"未完成缺陷\s*[:：]?\s*(.*?)(?=已完成缺陷)", re.DOTALL)
_E2_FINISHED_BLOCK_RE = re.compile(r"已完成缺陷\s*[:：]?\s*(.*)", re.DOTALL)
# xlsx 文件名常见「v2.4.4版本用例（…）」：标题里只保留【v2.4.4版本】，去掉用例说明与重复「版本」
_TITLE_USE_CASE_RE = re.compile(r"用例[（(][^）)]*[）)]")
# 标题行「【Yaahlan2.4.6版本】」：从文件名中取 x.y.z
# 不用尾随 \\b：「2.4.6」后紧跟中文「版」时 ASCII \\b 不成立，会误走回退分支
_YAAHLAN_SEMVER_IN_STEM = re.compile(r"(?i)(v?)(\d+\.\d+\.\d+)")


def _title_prefix_from_xlsx_stem(stem: str) -> str:
    s = _TITLE_USE_CASE_RE.sub("", stem.strip()).strip()
    if s.endswith("版本"):
        s = s[:-2].strip()
    return s if s else stem.strip()


def _yaahlan_version_s2_from_stem(stem: str) -> str:
    """内网/外网报告标题中 s2 文案：Yaahlan + 主版本号，与「【Yaahlan2.4.6版本】」格式一致。"""
    s = stem.strip()
    m = _YAAHLAN_SEMVER_IN_STEM.search(s)
    if m:
        return f"Yaahlan{m.group(2)}"
    tp = _title_prefix_from_xlsx_stem(s)
    return f"Yaahlan{tp}" if tp else "Yaahlan"


def _normalize_subitem(s: str) -> str:
    s = s.strip()
    s = _SUBITEM_PREFIX_RE.sub("", s)
    return s.strip()


def _parse_sheet1_e2_urls(raw: str | None) -> tuple[str | None, str | None, str | None]:
    """从 Sheet1 I2 文本解析缺陷(TB)、版本用例、回归用例三类 URL。"""
    if raw is None:
        return (None, None, None)
    text = str(raw).strip()
    if not text:
        return (None, None, None)

    def _m_url(pat: re.Pattern[str]) -> str | None:
        m = pat.search(text)
        return m.group(1).strip() if m else None

    return (
        _m_url(_E2_DEFECT_URL_RE),
        _m_url(_E2_VERSION_CASE_URL_RE),
        _m_url(_E2_REGRESSION_URL_RE),
    )


def _e2_defect_levels_from_block(block: str) -> dict[str, int]:
    d = {"阻碍": 0, "严重": 0, "一般": 0, "轻微": 0}
    for name, num_s in _E2_DEFECT_LEVEL_LINE_RE.findall(block):
        d[name] = int(num_s)
    return d


def _defect_levels_dict_empty() -> dict[str, int]:
    return {"阻碍": 0, "严重": 0, "一般": 0, "轻微": 0}


def _parse_i2_unfinished_finished_dicts(raw: str | None) -> tuple[dict[str, int], dict[str, int]]:
    """从 I2 原文解析「未完成 / 已完成」两段的各级数量（与表内用 `；` 连写或分行均可）。"""
    e = _defect_levels_dict_empty()
    if raw is None:
        return (e.copy(), e.copy())
    text = str(raw).strip()
    if not text:
        return (e.copy(), e.copy())
    unf_block_m = _E2_UNFINISHED_BLOCK_RE.search(text)
    fin_block_m = _E2_FINISHED_BLOCK_RE.search(text)
    unf_block = unf_block_m.group(1) if unf_block_m else ""
    fin_block = fin_block_m.group(1) if fin_block_m else ""
    return (
        _e2_defect_levels_from_block(unf_block),
        _e2_defect_levels_from_block(fin_block),
    )


def _format_i2_defect_level_row(levels: dict[str, int] | None) -> str:
    """与 Excel 调整版一致：一行内「阻碍：0；严重：0；一般：1；轻微：2」。"""
    d = levels if levels is not None else _defect_levels_dict_empty()
    return "；".join(f"{k}：{d.get(k, 0)}" for k in ("阻碍", "严重", "一般", "轻微"))


def format_sheet1_i2_text(
    *,
    defect_tb_url: str | None,
    version_case_url: str | None,
    regression_case_url: str | None,
    unf: dict[str, int] | None = None,
    fin: dict[str, int] | None = None,
) -> str:
    """生成与当前版本用例表 I2 相同结构的多行文本（缺陷/版本/回归 URL + 未完成&已完成统计）。"""
    lines: list[str] = [
        f"缺陷链接：{defect_tb_url or ''}",
        f"版本用例链接：{version_case_url or ''}",
        f"回归用例链接：{regression_case_url or ''}",
        f"未完成缺陷：{_format_i2_defect_level_row(unf)}",
        f"已完成缺陷：{_format_i2_defect_level_row(fin)}",
    ]
    return "\n".join(lines)


def reformat_raw_i2_like_excel(i2_raw: str | None) -> str:
    """从 I2 原文解析链接与数字后，按 `format_sheet1_i2_text` 规则重排为统一格式。"""
    tb, ver, reg = _parse_sheet1_e2_urls(i2_raw)
    unf, fin = _parse_i2_unfinished_finished_dicts(i2_raw)
    return format_sheet1_i2_text(
        defect_tb_url=tb,
        version_case_url=ver,
        regression_case_url=reg,
        unf=unf,
        fin=fin,
    )


def _parse_sheet1_e2_defect_stats(raw: str | None) -> tuple[int, int, int, int, int, int, int, str] | None:
    """解析 I2 中未完成/已完成缺陷统计，返回
    (发现缺陷总数, 1~4级数量, 已修复, 未修复, 缺陷修复率百分比字符串)。
    缺陷修复率按「已修复/发现缺陷×100」保留两位小数。"""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None

    unf, fin = _parse_i2_unfinished_finished_dicts(raw)

    unf_block_m = _E2_UNFINISHED_BLOCK_RE.search(text)
    fin_block_m = _E2_FINISHED_BLOCK_RE.search(text)

    unfixed = sum(unf.values())
    fixed = sum(fin.values())
    total = unfixed + fixed
    if total == 0 and not (unf_block_m or fin_block_m):
        return None

    l1 = unf["阻碍"] + fin["阻碍"]
    l2 = unf["严重"] + fin["严重"]
    l3 = unf["一般"] + fin["一般"]
    l4 = unf["轻微"] + fin["轻微"]

    if total == 0:
        rate_s = "0.00"
    else:
        rate_s = f"{(fixed / total) * 100:.2f}"

    return (total, l1, l2, l3, l4, fixed, unfixed, rate_s)


def _inject_defect_stats(html_out: str, stats: tuple[int, int, int, int, int, int, int, str] | None) -> str:
    """将 I2 缺陷统计写入模板中「发现缺陷」列表项（位于 tb 地址之前）。"""
    if stats is None:
        return html_out
    total, l1, l2, l3, l4, fixed, unfixed, rate_s = stats

    start = html_out.find('<span class="s1">发现缺陷</span>')
    tb = '<span class="s2">tb</span>'
    end = html_out.find(tb, start)
    if start < 0 or end < 0 or start >= end:
        return html_out

    region = html_out[start:end]
    vals = [str(total), str(l1), str(l2), str(l3), str(l4), str(fixed), str(unfixed)]
    new_region = region
    for v in vals:
        if '<span class="s5">x</span>' not in new_region:
            return html_out
        new_region = new_region.replace('<span class="s5">x</span>', f'<span class="s5">{v}</span>', 1)
    if "缺陷修复率x" not in new_region:
        return html_out
    new_region = new_region.replace("缺陷修复率x", f"缺陷修复率{rate_s}", 1)

    return html_out[:start] + new_region + html_out[end:]


def _html_a_href(url: str, inner_text: str, *, a_class: str = "") -> str:
    esc_url = html.escape(url, quote=True)
    esc_inner = html.escape(inner_text)
    cls = f' class="{a_class}"' if a_class else ""
    return f'<a href="{esc_url}"{cls}>{esc_inner}</a>'


def _inject_report_links(
    html_out: str,
    *,
    defect_tb_url: str | None,
    version_case_url: str | None,
    regression_case_url: str | None,
) -> str:
    """将 I2 中的链接写入「用例地址」「tb 地址」及覆盖表格中两处详情占位。"""
    out = html_out

    if version_case_url:
        frag = _html_a_href(version_case_url, version_case_url, a_class="s7")
        out = out.replace(
            '<span class="s7">用例地址：x</span>',
            f'<span class="s7">用例地址：{frag}</span>',
            1,
        )

    if defect_tb_url:
        frag = _html_a_href(defect_tb_url, defect_tb_url, a_class="s8")
        out = out.replace(
            '<span class="s2">tb</span><span class="s8">地址：x</span>',
            f'<span class="s2">tb</span><span class="s8">地址：{frag}</span>',
            1,
        )

    def _replace_first_between(start_m: str, end_m: str, old: str, new: str) -> None:
        nonlocal out
        i = out.find(start_m)
        if i < 0:
            return
        j = out.find(end_m, i + len(start_m))
        if j < 0:
            return
        region = out[i:j]
        if old not in region:
            return
        out = out[:i] + region.replace(old, new, 1) + out[j:]

    if version_case_url:
        vfrag = _html_a_href(version_case_url, version_case_url, a_class="s3")
        new_cell = f'<span class="s3">{vfrag}</span>'
        _replace_first_between("新功能覆盖", "原功能回归", '<span class="s3">x</span>', new_cell)

    if regression_case_url:
        rfrag = _html_a_href(regression_case_url, regression_case_url, a_class="s3")
        new_cell = f'<span class="s3">{rfrag}</span>'
        _replace_first_between("原功能回归", "分区测试", '<span class="s3">x</span>', new_cell)

    return out


def _open_html_default_browser(html_path: Path) -> bool:
    """用系统默认浏览器/应用打开本地 HTML；macOS 用 `open`，比 webbrowser 在部分终端里更稳。"""
    path = html_path.resolve()
    if not path.is_file():
        return False
    system = platform.system()
    try:
        if system == "Darwin":
            r = subprocess.run(
                ["open", str(path)],
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
            return r.returncode == 0
        if system == "Windows":
            os.startfile(str(path))  # type: ignore[attr-defined]
            return True
        r = subprocess.run(
            ["xdg-open", str(path)],
            check=False,
            capture_output=True,
            text=True,
            timeout=30,
        )
        if r.returncode == 0:
            return True
    except (OSError, subprocess.TimeoutExpired) as e:
        print(f"系统命令打开失败（{e}），尝试 webbrowser…", file=sys.stderr)
    try:
        return bool(webbrowser.open(path.as_uri()))
    except OSError as e:
        print(f"webbrowser 打开失败：{e}", file=sys.stderr)
        return False


# 测试覆盖表：内嵌模板用「相邻单元格一边透明」省线，但 Gmail/Outlook/部分企业邮会丢格线或最右侧线；统一实线 + 可伸缩宽度后黏贴到邮件里更稳
# 与模板中 p.p3 / p.p4 等一致：10.5pt 宋体正文色；子元素 span.s2、span.s5 等仍用各自类里的 Helvetica
_COVERAGE_TABLE_FONT = (
    "font: 10.5pt/1.35 'Songti SC', 'PingFang SC', 'Hiragino Sans GB', 'STSong', 'SimSun', 'Microsoft YaHei', "
    "serif; color: #0d0e11; "
)
_COVERAGE_TABLE_CSS_FOR_EMAIL = f"""    table.t1 {{ width: 100%; max-width: 746px; border-collapse: collapse; table-layout: fixed; border: 1px solid #858889; -webkit-text-size-adjust: 100%; mso-fareast-font-family: 'Songti SC', SimSun, serif; {_COVERAGE_TABLE_FONT} }}
    td.td1, td.td2, td.td3, td.td4, td.td5, td.td6, td.td7, td.td8 {{ border: 1px solid #858889; padding: 4.0px 7.0px 4.0px 7.0px; vertical-align: top; box-sizing: border-box; font: inherit; }}
    td.td1, td.td5 {{ width: 46.4px; }}
    td.td2, td.td6 {{ width: 118.5px; }}
    td.td3, td.td7 {{ width: 37.6px; }}
    td.td4, td.td8 {{ width: 482.4px; min-width: 0; word-wrap: break-word; }}
    """
_COVERAGE_TABLE_TAG_LEGACY = '<table width="746.0" cellspacing="0" cellpadding="0" class="t1">'
_COVERAGE_TABLE_TAG_FOR_EMAIL = (
    '<table class="t1" width="100%" border="1" bordercolor="#858889" cellspacing="0" '
    "cellpadding=\"0\" style=\"max-width:746px;border:1px solid #858889;border-collapse:collapse;"
    "mso-fareast-font-family:'Songti SC',SimSun,serif;"
    f"{_COVERAGE_TABLE_FONT.rstrip()}"
    '">'
)


def _coverage_table_css_block_from_template() -> str | None:
    """从内嵌模板取出「table.t1 … td.td8」这段旧 CSS，供与生成结果做 str.replace；模板改版式时若结构仍含 table.t1 / ul.ul2 可自动适配。"""
    text = _report_template_html()
    i = text.find("    table.t1 {")
    j = text.find("    ul.ul2 {", i)
    if i < 0 or j < 0 or j <= i:
        return None
    return text[i:j]


def _inject_email_table_style_fallback(html: str) -> str:
    """未匹配到内嵌模版权重子串时，在 </head> 前追加高优先级规则（旧版已生成 HTML 重跑时可能用到）。"""
    font_s = _COVERAGE_TABLE_FONT.rstrip()
    inj = (
        '  <style type="text/css">\n'
        "  /* count.py: 邮件格线 + 与正文一致的中文字体（见 _COVERAGE_TABLE_FONT） */\n"
        f"  table.t1 {{ width: 100% !important; max-width: 746px !important; border-collapse: collapse !important; border: 1px solid #858889 !important; mso-fareast-font-family: 'Songti SC', SimSun, serif !important; {font_s} }}\n"
        "  table.t1 td.td1, table.t1 td.td2, table.t1 td.td3, table.t1 td.td4, table.t1 td.td5, table.t1 td.td6, table.t1 td.td7, table.t1 td.td8 { border: 1px solid #858889 !important; padding: 4.0px 7.0px; vertical-align: top; box-sizing: border-box; font: inherit !important; }\n"
        "  </style>\n"
    )
    if "</head>" in html:
        return html.replace("</head>", inj + "</head>", 1)
    return html


def _harden_coverage_table_for_email(html: str) -> str:
    """为抄送邮件等场景强化「一、测试覆盖情况」表格外框与格线，避免最右侧/透明边在部分客户端不显示。"""
    out = html
    legacy = _coverage_table_css_block_from_template()
    if legacy and legacy in out:
        out = out.replace(legacy, _COVERAGE_TABLE_CSS_FOR_EMAIL, 1)
    else:
        out = _inject_email_table_style_fallback(out)
    if _COVERAGE_TABLE_TAG_LEGACY in out:
        out = out.replace(_COVERAGE_TABLE_TAG_LEGACY, _COVERAGE_TABLE_TAG_FOR_EMAIL, 1)
    return out


# 外网「功能开放前 checklist」表：仅写一份 *外网测试总结.html*，线色 #858 与内网表邮件版一致，便于发邮件/复制
_EXTERNAL_CHECKLIST_TABLE_FONT = (
    "font: 14px/1.35 'Songti SC', 'PingFang SC', 'Hiragino Sans GB', 'STSong', 'SimSun', 'Microsoft YaHei', "
    "serif; color: #0d0e11; "
)
_EXTERNAL_CHECKLIST_TABLE_BORDER = "#858889"  # 同内网表 _harden
_EXTERNAL_CHECKLIST_TABLE_TAG_LEGACY = '<table width="362.0" cellspacing="0" cellpadding="0" class="t1">'


def _external_checklist_table_css_hardened(table_border: str) -> str:
    return f"""    table.t1 {{ width: 100%; max-width: 362px; border-collapse: collapse; table-layout: fixed; border: 1px solid {table_border}; -webkit-text-size-adjust: 100%; mso-fareast-font-family: 'Songti SC', SimSun, serif; {_EXTERNAL_CHECKLIST_TABLE_FONT} }}
    td.td1, td.td2, td.td3, td.td4 {{ border: 1px solid {table_border}; padding: 4.0px 8.0px 4.0px 8.0px; vertical-align: top; box-sizing: border-box; font: inherit; }}
    td.td1, td.td3 {{ width: 185.6px; }}
    td.td2, td.td4 {{ width: 141.4px; min-width: 0; word-wrap: break-word; }}
    """


def _external_checklist_table_tag_hardened(table_border: str) -> str:
    return (
        f'<table class="t1" width="100%" border="1" bordercolor="{table_border}" cellspacing="0" '
        f'cellpadding="0" style="max-width:362px;border:1px solid {table_border};border-collapse:collapse;'
        "mso-fareast-font-family:'Songti SC',SimSun,serif;"
        f"{_EXTERNAL_CHECKLIST_TABLE_FONT.rstrip()}"
        '">'
    )


def _external_checklist_table_css_block_from_template() -> str | None:
    """从外网模板取「功能开放前 checklist」table.t1 ～ td4 的 CSS 片段（后接 ul.ul1）。"""
    text = _external_report_template_html()
    i = text.find("    table.t1 {")
    j = text.find("    ul.ul1 {", i)
    if i < 0 or j < 0 or j <= i:
        return None
    return text[i:j]


def _inject_external_email_table_style_fallback(html: str, table_border: str) -> str:
    fnt = _EXTERNAL_CHECKLIST_TABLE_FONT.rstrip()
    b = table_border
    inj = (
        '  <style type="text/css">\n'
        "  /* count.py: 外网 checklist 表格线（不依赖 border-color: transparent） */\n"
        f"  table.t1 {{ width: 100% !important; max-width: 362px !important; border-collapse: collapse !important; border: 1px solid {b} !important; mso-fareast-font-family: 'Songti SC', SimSun, serif !important; {fnt} }}\n"
        f"  table.t1 td.td1, table.t1 td.td2, table.t1 td.td3, table.t1 td.td4 {{ border: 1px solid {b} !important; padding: 4.0px 8.0px; vertical-align: top; box-sizing: border-box; font: inherit !important; }}\n"
        "  </style>\n"
    )
    if "</head>" in html:
        return html.replace("</head>", inj + "</head>", 1)
    return html


def _harden_external_checklist_table(html: str, table_border: str) -> str:
    """将外网「功能开放前 checklist」的透明/断裂边线改为实线，颜色由 `table_border` 指定（默认用 `_EXTERNAL_CHECKLIST_TABLE_BORDER` 与内网表一致）。"""
    out = html
    legacy = _external_checklist_table_css_block_from_template()
    if legacy and legacy in out:
        out = out.replace(legacy, _external_checklist_table_css_hardened(table_border), 1)
    else:
        out = _inject_external_email_table_style_fallback(out, table_border)
    if _EXTERNAL_CHECKLIST_TABLE_TAG_LEGACY in out:
        out = out.replace(_EXTERNAL_CHECKLIST_TABLE_TAG_LEGACY, _external_checklist_table_tag_hardened(table_border), 1)
    return out


# 内嵌报告 HTML 骨架（样式 + 覆盖表格 + 尾部占位），zlib+base64；改版式需同步更新此串或从 1.html 重新生成
_REPORT_TEMPLATE_ZLIB_B64 = "eNrtW21P21YU/t5fcZdpqlQpcZw4L4SAtEHVVuqbRraqH53YgFVjZ/ZNoUKVWCteUiClE6O0ZaWMDlhboHRqCSFZpf2VxXbyib+wa9+8QR2aBCesLU0V29f3POec+zz3+DpXBL/qvtIVun71LOiHAzy4+sN3Fy90AZudIK65uwiiO9QNzocuXQSUw0kSxNnLNmDrhzAaIIjBwUHHoNshSn1E6HtCt6YIGUpcBDoYyNg6TwX1Nv3A0kznKQCCAyykgW5tZ3+KcTc7bF2iAFkB2kO3oqwNRPBVhw2yQ9AAbAeRflqSWdgRg712v+1wlB54i2dNsSKyjG0hB3m2M0jgYwlNoAfYDts5VmAlGopShXmXGBFpPALXJA6yku2gldHjR1aSOVGoMHR5/R4HhXvLemAAosAOxANA1BElwfAALfVxQgA4Hc7o0Iff7aAXwQYA6XR4ohCc7hGFPsiBnq7TaIBEXpQC4Gsn42RJ8nYR1FU7qAtd1Qbqbkak1BFASyAeK0C8dYOcZ/mbLOQiNLjMxtjqOfqOjjzACfZ+luvr1ztRukkJ3d+cuIkzIJtMgBgPlImHuaW1/NKOur6szo0r2ytael7Z3MlnZvbSj/FdwHByBOR353Mbz7XUirKwBijdvzKTyL3d1l7NgjOEARvjHTEeaZ7nZGg3ZoZdnxkBAwDlWcykHURphuGEPjvP9qLYqcqUSWdFzsaHMtLcb+Ly6236jLNzAsPqI2DHbXWrlCQPOvT6rXBYduCyjkWyTf+Uod3NEQjPOXiuzfKaYMCSlIW47SBMR270SWJMYOzFW73GP+xSjtKCQ0YU64j2G6wkIE4DQBAFtkZbF7Y9ZAQbhnabhVXZAQ1VTUCe5sXoPQAd4gZY+Qh4voOhVpLbMKrfYlRIh3nWAZFuBjkG9geAj/JiZYZFiWEl3YinozIqbsWzgiHjgEzZDFlRFVZGUQwAWeQ5ptRY6Eoa+v/gu9KjEabf4/f726odS3UrAIxHCfAZ3xXnFXG6SnGSpB+NW6OBOusJFEq0gEiTUA2tK1h3KVi3z+H9f8dKlQXgdx1BAS0J1tOIWp21qbUiJksU621Isc7aBvawYBsdXF8jqj3GeP0NKfd4AjbWmy6T9WaEkyK8UZCDhNGOXlUJ/K4aDIvMrc5Tw8OhC6GLZ2/fDkZBhKdlucMWddk6g/rTpNggk6gh3JlNjvw78rP6djK3+WtuZUx7MqfeHVXG3gaJsH7bOBC6HTpEkQPj0QGMgemwGY8N9O7I8jzqEkE5ddgK14Uc8TX2CMnCuywOUs8xCCV8op8y4CbNc30C6ilGy1YMaSv2Qb3KCbnNE8KpaFu7ytPJKkkUoAjIfNy5y9w5ZeL8n2eNOnHXl+H8pjKzgsmyIEOqLue5zRUkkBrdojOpDqI9dUWi3FvM3c1gui0YBq+5c08VDuZeY/+W0eCrPXkkgEa9+GuUM3I71HxmrZ1F9TGoJBYxg8qTp0rml+Yx6DlhsEkMTowpUylcAPbSE0o6paS38o9ntZVd5fkqurWXjjePVq9JUITFpPpMSC26CEvHUXNH08rGjjqyelxldyGFHus4Cjt3pYek7C5va4svevRazLKZl/zCiLp1J5ucVt/Flfv3tL92td1FnPdeeiqbGcMEqPFJfUgWU+rCK/wkUjZ3lNQs7v95T/99WqAFRhI5xmcnT+TwZcpBi0+grAtyyL+cz2YSuOlTX5uZqm77jbKwVrkCzr0f19Yms8l1zDtSRW618Fqnza5l/55U58bVpeUvSAPKxu/qxLb2chO35h/NKBPvTsTQUjFUW7lNj2up41pDKNPJXCajbMRzy6P6yOA6aVyeqKOl6qi6yHyX3Z1rni6q/awysmrl7wl1COHzeW1oSfGvJpv0lpZ5cMLfJ8vfzh1tPW7lu+UJhS2mML+cUBMn/H26U3Ajroyufen8oSPeoUIn+m4XOikD+z9KTw17bqmp0p6bGl/NLU0dvudWBiQb3goZHu6+fvnbSxe69m0Kkm7TdPQPKP0vwuzrhwYYLxaVhdfKbyN76cf7fm0Oxvhi1xhf2PfjuWITz7Ud9EuV+flgBXv/gZZ4raVT+Ufbpn08FT91HzTOJl+gBa6+pEyuVzMmqxmjl5RGHCJyq5m5WufK3TpXlMWudMa232TfbyjPpxu0VxdeHMkeyw1DaInxoWow31SDQRimt1BNgGHTO6i0mEwmntMLUYy3dk5lk9PK/U1UMXDlQR5rdFjZVNwbbzgK9eGzfW712ms4tjxfJCdt8Q+0Psk/+tPCZEmzbbePZEUUHy74rwz+A+B8ri4="


@functools.lru_cache(maxsize=1)
def _report_template_html() -> str:
    return zlib.decompress(base64.b64decode(_REPORT_TEMPLATE_ZLIB_B64.encode("ascii"))).decode(
        "utf-8"
    )


def _build_dynamic_section_html(
    *,
    new_need_count: int,
    optimize_need_count: int,
    new_titles: list[str],
    tech_opt_items: list[str],
) -> str:
    _p8 = _p8_blank_paragraph()
    parts: list[str] = [
        _p8,
        '<ul class="ul1">\n',
        '  <li class="li9"><span class="s4"></span><span class="s1">本次版本共</span>'
        f'<span class="s5">{new_need_count}</span><span class="s1">个新需求及</span>'
        f'<span class="s5">{optimize_need_count}</span><span class="s1">个优化需求</span></li>\n',
        "</ul>\n",
        _p8,
    ]
    for i, title in enumerate(new_titles, start=1):
        t = html.escape(title)
        parts.append(
            f'<p class="p10"><span class="s5">{i}.</span><span class="s1">{t}</span></p>\n'
        )
    if tech_opt_items:
        n = len(new_titles) + 1
        parts.append(
            f'<p class="p10"><span class="s5">{n}.</span><span class="s1">技术优化需求：</span></p>\n'
        )
        for j, item in enumerate(tech_opt_items, start=1):
            parts.append(
                f'<p class="p11"><span class="s5">{j}</span><span class="s1">）{html.escape(item)}</span></p>\n'
            )
    # 与模板中 p12 空白段一致，便于衔接「用例地址」所在 p13
    parts.append('<p class="p12"><span class="s1">\u00a0</span></p>\n')
    return "".join(parts)


def _p8_blank_paragraph() -> str:
    """与动态区、模板中 p8 一致，占一行高以形成空行。"""
    return '<p class="p8"><span class="s1">\u00a0</span></p>\n'


def _insert_internal_report_title_spacers(html: str) -> str:
    """标题行 与「一、测试覆盖情况」之间、「一、测试覆盖情况」与覆盖表格之间各加一空行（与 p8 样式一致）。"""
    blank = _p8_blank_paragraph()
    a = "内网测试已完成，结果总结如下：</span></p>\n<p class=\"p2\"><span class=\"s1\"><b>一、测试覆盖情况</b>"
    b = "内网测试已完成，结果总结如下：</span></p>\n" + blank + "<p class=\"p2\"><span class=\"s1\"><b>一、测试覆盖情况</b>"
    if a in html:
        html = html.replace(a, b, 1)
    c = "<b>一、测试覆盖情况</b><b></b></span></p>\n<table "
    d = "<b>一、测试覆盖情况</b><b></b></span></p>\n" + blank + "<table "
    if c in html:
        html = html.replace(c, d, 1)
    return html


def _insert_external_checklist_title_spacer(html: str) -> str:
    """在「功能开放前 checklist」与下方表格之间加一空行（与内网 p8 空行一致）。"""
    a = (
        '<p class="p6"><span class="s8"><b>功能开放前</b></span><span class="s1">checklist</span></p>\n<table'
    )
    b = (
        '<p class="p6"><span class="s8"><b>功能开放前</b></span><span class="s1">checklist</span></p>\n'
        + _p8_blank_paragraph()
        + "<table"
    )
    if a in html:
        return html.replace(a, b, 1)
    return html


def _render_full_report_html(
    *,
    title_prefix: str,
    new_need_count: int,
    optimize_need_count: int,
    new_titles: list[str],
    tech_opt_items: list[str],
) -> str:
    t = html.escape(title_prefix)
    title_line = (
        f'<p class="p1"><span class="s1"><b>【</b></span><span class="s2">{t}</span>'
        f'<span class="s1"><b>版本】</b>内网测试已完成，结果总结如下：</span></p>\n'
    )
    dynamic = _build_dynamic_section_html(
        new_need_count=new_need_count,
        optimize_need_count=optimize_need_count,
        new_titles=new_titles,
        tech_opt_items=tech_opt_items,
    )
    tpl = _report_template_html()
    out = tpl.replace("{{TITLE}}", title_line, 1).replace("{{DYNAMIC}}", dynamic, 1)
    return _insert_internal_report_title_spacers(out)


# 外网测试总结 HTML（与「*_外网测试总结.html」版式一致）；改版式需同步更新此串或从样例 HTML 重新生成
_EXTERNAL_REPORT_TEMPLATE_ZLIB_B64 = "eNrFWF1T20YUfc+v2KoPebJk2cY2xvZDjSfJlJIUu2H6xMjSYlSEpEprHIZhxkkGGkhoSRvIlDAJnTZMO9MC7UxSUiCdyV8JkuEpf6GrXX8IIzu2cYkfpNXq3nvOnrt7tev4R4PXU9kvb6TBBJpSwI0vPhm6lgKMj+NGgymOG8wOgqvZz4ZAiPXzHJceZgAzgZAe47hiscgWg6xm5LnsCOd4hzgTGbKIWAlJTPJS3OlzblCQkpcAiE9BJADH2we/LsjTCSalqQiqyJed0SEDRPqUYBC8hUjAASBOCIYJUaKAxn1RpnWUDJpRoGcs0TSpL5KRApNxjt5r0VRhCiaYK1CFhoA0w+We0kRNoAqMGjKCBtPoRSxuQsOUNdXlGAhH+9gQtTYdYgBhYg18ANBZnQezU4KRl9UY8LN+/dbZ6wAYx2FjgA+RvssZTc0jGWRSl7FAmqIZMfCxX/JDnh8AviLMTcrI5+D4cDq0SVh7O1eFDLQBSQIIipzHRl8VTCSPz/ScR7g3PK5CZRoiWRTAMCzAbslEzkkmyAY9yTQBJ78aeLQ34O6MtIXb33vcrsTn/b2fkm0JwPM9Ac7KU9BsDTkAcoI4mTe0gir5qiKNkx9lo8isIgc7qQaRzodMQELnAuk8wQS07yJAXSM1dUFlTZxeJ6hvEhqqrOZjQNVU+N5UUN8A9W1RaLoOjdPcll3IRSHSWwp9LSg00dcRwUPjcIcat4zumjY0eqSrLNTmS3trj0JFG6HcM7BroftrUZt9IroO7ZTNs87UAgk5BbIIL4CiLKGJGAiGA3SV5TRDgoYTVhF0EytfbVUcJRZJdTceb2TCLjeym4kBU1NkqdZZtSWynbm6IV1ZaXYfALogSWQ4NA9RcnW1XUQDdaIhng11TdTfCVFkCCrOgIG3eh2RDXalqr89VV2keqJsqCtl/e0p24psF+oWFLag4DmryCai9HzOfjsGJNkUXSYBDxNRNkSFzP04R/rxmYWjh5Z4TpNm8E0HoiKYZoLReSYZd9ZetcN0OnLJt6WVOJfDJwvnXYNFgEnOzt5Mj2SuXR8eywTm5jytaJzy4j174/e3pYdONOuXtfLhQ/vF/eOdVevvv6ztB/a9lXcHD8r7P9hPN+zSPm5YW3eO9u6/O1ivBuX0U3wD3nytpWfHdw9paELc4V7nT4IUlKoTlpaeZBS52oV3Ko2RnY4mI7PXdk82Svafdyji8dZC+cla+dGvR69d1Gdnh9OjYyPpz8eGrg1/6sikyE4qCkpvyVhPnlqH33u+xQc2cQKKk84caebejPxI+spIOkOS7ObfQDTUSLSvOdGjvWXrux0smX133lp44U5yJbJblUDlSHkKrq9dXcI4SY83G+MT7c8xAjxry8+en/z87cmPv10g/VOvIkzyjffAKtf6Ygk3wkVdi8U6KNmPXluLy00XOu8xe8hKIp9iQApkgiGfYQaIUFGwiYhrW4KpPFdqHX2mQVFlsiNaiYhCyKANpymBaXIwwZaaXveSeKZqg63qI4wwbXKueHJIej9WwBsr6l16yksv7dLtpiL2M0muRYE83tnCq6FJyWokjVtGB5oFOxvH7V3rn63yH2vl1edt8mkFHmof3L1QL3rUtIDbq7v28ra9N19eWbjYseOKUoX9QAoc7S3aa6/szT1r59XJ/HL5cPv/U6DfK/sUubR+/O83H0wEa3vdfvzyaH/fWvrpgoff3uTH98rejSPV9/Qmzu8R+M1mky0Uf3bP5z9jzlXR6J/e/wEGSoi7"


@functools.lru_cache(maxsize=1)
def _external_report_template_html() -> str:
    return zlib.decompress(base64.b64decode(_EXTERNAL_REPORT_TEMPLATE_ZLIB_B64.encode("ascii"))).decode(
        "utf-8"
    )


def _render_external_report_html(
    *,
    version_s2: str,
    version_case_url: str | None,
    regression_case_url: str | None,
) -> str:
    """外网报告：标题版本号与内网同源（xlsx 文件名推导）；I2 版本用例/回归链接插入对应列表项。"""
    tpl = _external_report_template_html()
    html_out = tpl.replace("{{VERSION_S2}}", html.escape(version_s2))
    new_frag = (
        " " + _html_a_href(version_case_url, version_case_url, a_class="s1") if version_case_url else ""
    )
    reg_frag = (
        " " + _html_a_href(regression_case_url, regression_case_url, a_class="s1")
        if regression_case_url
        else ""
    )
    html_out = html_out.replace("{{NEW_REQ_LINK}}", new_frag).replace("{{REGRESSION_LINK}}", reg_frag)
    # 原模板两处独立 ul 会在浏览器里拉出空行；合并为同一 ul 下的连续 li
    html_out = html_out.replace(
        '</li>\n</ul>\n<ul class="ul1">\n  <li class="li3"><span class="s3"></span><span class="s1">回归</span>',
        '</li>\n  <li class="li3"><span class="s3"></span><span class="s1">回归</span>',
        1,
    )
    return _insert_external_checklist_title_spacer(html_out)


def _report_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _project_root() -> Path:
    return _report_root().parent


def _find_venv_python() -> str | None:
    for candidate in (
        _report_root() / ".venv" / "bin" / "python",
        _project_root() / ".venv-xlsx" / "bin" / "python",
        _project_root() / ".venv" / "bin" / "python",
    ):
        if candidate.is_file() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def _resolve_xlsx_path_and_print_i2_flags(argv: list[str], base_dir: Path) -> tuple[str, bool] | None:
    """从 argv 得到 xlsx 路径与是否只输出 I2 文本。失败时打印错误并返回 None。"""
    want_print_i2 = "--print-i2" in argv
    rest = [a for a in argv[1:] if a != "--print-i2"]
    if rest:
        xlsx_path = rest[0]
        if not os.path.isabs(xlsx_path):
            xlsx_path = str((base_dir / xlsx_path).resolve())
        if not os.path.exists(xlsx_path):
            print(f"找不到文件：{xlsx_path}", file=sys.stderr)
            return None
        return (xlsx_path, want_print_i2)
    xlsx_candidates = sorted(base_dir.glob("*.xlsx"))
    if not xlsx_candidates:
        print(
            f"在目录 {base_dir} 下未找到任何 .xlsx 文件；请指定路径，例如：\n"
            "  python3 Report/report_execute.py /path/to/v2.4.4版本用例.xlsx",
            file=sys.stderr,
        )
        return None
    return (str(xlsx_candidates[0]), want_print_i2)


def main() -> int:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        venv_python = _find_venv_python()
        if venv_python:
            entry = str(_report_root() / "report_execute.py")
            os.execv(venv_python, [venv_python, entry, *sys.argv[1:]])
        print(
            "未安装 openpyxl。请先执行：\n"
            "  python3 -m venv Report/.venv\n"
            "  source Report/.venv/bin/activate\n"
            "  python -m pip install -r Report/requirements.txt",
            file=sys.stderr,
        )
        return 2

    base_dir = Path.cwd()
    resolved = _resolve_xlsx_path_and_print_i2_flags(sys.argv, base_dir)
    if resolved is None:
        return 2
    xlsx_path, want_print_i2 = resolved

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except PermissionError:
        print(f"无法读取文件（可能正在被 Excel 占用）：{xlsx_path}", file=sys.stderr)
        return 2
    except OSError as e:
        print(f"读取 Excel 失败：{e}", file=sys.stderr)
        return 2

    # 某些 xlsx 首个 sheet 可能是 hidden 的「AI生成」草稿（I2 为空），
    # 真正的汇总数据在后续 visible sheet 上。若无 Sheet1，优先选 I2 非空的 visible sheet。
    def _pick_i2_sheet_name() -> str:
        if "Sheet1" in wb.sheetnames:
            return "Sheet1"
        visible_names = [
            n for n in wb.sheetnames
            if getattr(wb[n], "sheet_state", "visible") == "visible"
        ]
        for n in visible_names:
            if wb[n]["I2"].value is not None:
                return n
        if visible_names:
            return visible_names[0]
        return wb.sheetnames[0]

    sheet1_name = _pick_i2_sheet_name()
    i2_raw = wb[sheet1_name]["I2"].value

    if want_print_i2:
        print(reformat_raw_i2_like_excel(i2_raw))
        return 0

    lines: list[str] = []
    new_titles: list[str] = []
    tech_opt_items: list[str] = []
    new_need_count = 0
    optimize_need_count = 0
    idx = 1
    for name in wb.sheetnames:
        ws = wb[name]
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        v = ws["D2"].value
        if v is None:
            continue
        raw = str(v)
        if not raw.strip():
            continue

        parts = [p.strip() for p in raw.splitlines()]
        parts = [p for p in parts if p]
        if not parts:
            continue

        first = parts[0]
        rest = parts[1:]

        lines.append(f"{idx}. {first}")
        if _is_tech_opt_need(first) and rest:
            normalized_rest = []
            for item in rest:
                n = _normalize_subitem(item)
                if n:
                    normalized_rest.append(n)
            for j, item in enumerate(normalized_rest, start=1):
                lines.append(f" {j}）{item}")
            optimize_need_count += len(normalized_rest)
            tech_opt_items.extend(normalized_rest)
        elif not _is_tech_opt_need(first):
            new_need_count += 1
            new_titles.append(first)
        idx += 1

    print(f"本次版本共{new_need_count}个新需求及{optimize_need_count}个优化需求")
    print("\n".join(lines))

    xlsx_file = Path(xlsx_path)
    title_s2_version = _yaahlan_version_s2_from_stem(xlsx_file.stem)
    out_html = _render_full_report_html(
        title_prefix=title_s2_version,
        new_need_count=new_need_count,
        optimize_need_count=optimize_need_count,
        new_titles=new_titles,
        tech_opt_items=tech_opt_items,
    )
    defect_stats = _parse_sheet1_e2_defect_stats(i2_raw)
    out_html = _inject_defect_stats(out_html, defect_stats)
    defect_tb_url, version_case_url, regression_case_url = _parse_sheet1_e2_urls(i2_raw)
    out_html = _inject_report_links(
        out_html,
        defect_tb_url=defect_tb_url,
        version_case_url=version_case_url,
        regression_case_url=regression_case_url,
    )
    out_html = _harden_coverage_table_for_email(out_html)

    html_path = xlsx_file.with_name(f"{xlsx_file.stem}_内网测试总结.html")
    try:
        html_path.write_text(out_html, encoding="utf-8")
    except OSError as e:
        print(f"写入 HTML 失败：{e}", file=sys.stderr)
        return 2
    print(f"已生成：{html_path}", file=sys.stderr)

    ext_src = _render_external_report_html(
        version_s2=title_s2_version,
        version_case_url=version_case_url,
        regression_case_url=regression_case_url,
    )
    ext_html = _harden_external_checklist_table(ext_src, _EXTERNAL_CHECKLIST_TABLE_BORDER)
    ext_path = xlsx_file.with_name(f"{xlsx_file.stem}_外网测试总结.html")
    try:
        ext_path.write_text(ext_html, encoding="utf-8")
    except OSError as e:
        print(f"写入外网 HTML 失败：{e}", file=sys.stderr)
        return 2
    print(
        f"已生成：{ext_path}（外网表线色同内网邮件，可直接复制到邮箱）",
        file=sys.stderr,
    )

    if os.environ.get("COUNT_NO_BROWSER", "").strip().lower() not in ("1", "true", "yes"):
        opened_ext = _open_html_default_browser(ext_path)
        opened_in = _open_html_default_browser(html_path)
        if opened_in and opened_ext:
            print("已用默认浏览器打开外网、内网报告（先外网后内网）。", file=sys.stderr)
        elif opened_in or opened_ext:
            print(
                "已打开部分报告；未能自动打开的文件请手动在访达/资源管理器中打开。",
                file=sys.stderr,
            )
        else:
            print("未能自动打开，请手动在访达/资源管理器中打开上述 HTML。", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
