"""从钉钉 Excel URL 拉取全量 sheet 数据并保存为本地 xlsx（保留行号，供 Report 使用）。"""

from __future__ import annotations

import asyncio
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook

_REPORT_ROOT = Path(__file__).resolve().parent.parent
_MCP_READ_DIR = (
    _REPORT_ROOT.parent
    / ".cursor/skills/testcase-to-excel/mcp_dingtalk_excel"
)
if str(_MCP_READ_DIR) not in sys.path:
    sys.path.insert(0, str(_MCP_READ_DIR))

from server_read import (  # type: ignore[import-untyped]
    API_BASE_URL,
    COMMON_HEADERS,
    DEFAULT_AEGIS_KEY,
    DEFAULT_AEGIS_SECRET,
    DEFAULT_TIMEOUT,
    DEFAULT_WORKID,
    MAX_CELLS_PER_REQUEST,
    clear_token_cache,
    extract_workbook_id_from_url,
    format_exception,
    format_http_error,
    getRangeData,
    getSheetList,
    getTokenAndOperatorId,
    is_invalid_auth_error,
    numberToColumnName,
)

_VERSION_STEM_RE = re.compile(r"(?:Yaahlan-)?(\d+\.\d+\.\d+)\s*版本")


class DingtalkExcelFetchError(RuntimeError):
    pass


def _raise_fetch(message: str) -> None:
    raise DingtalkExcelFetchError(message)


async def _fetch_sheet_values_raw(
    *,
    workbook_id: str,
    sheet_id: str,
    operator_id: str,
    access_token: str,
) -> list[list[Any]]:
    headers = {**COMMON_HEADERS, "x-acs-dingtalk-access-token": access_token}
    max_row = 10000
    max_column = "ZZ"
    column_count = 702

    async with httpx.AsyncClient(timeout=DEFAULT_TIMEOUT) as client:
        url_sheet_info = (
            f"{API_BASE_URL}/workbooks/{workbook_id}/sheets/{sheet_id}"
            f"?select=values&operatorId={operator_id}"
        )
        try:
            response = await client.get(url_sheet_info, headers=headers)
            response.raise_for_status()
            sheet_json = response.json()
            if sheet_json.get("rowCount"):
                max_row = sheet_json["rowCount"]
            if sheet_json.get("columnCount"):
                column_count = sheet_json["columnCount"]
                max_column = numberToColumnName(column_count)
        except httpx.HTTPError as e:
            _raise_fetch(f"获取 Sheet 维度失败:\n{format_http_error(e, url_sheet_info)}")

        total_cells = max_row * column_count
        if total_cells > MAX_CELLS_PER_REQUEST:
            rows_per_batch = max(1, MAX_CELLS_PER_REQUEST // column_count - 1)
            all_values: list[list[Any]] = []
            for start_row in range(1, max_row + 1, rows_per_batch):
                end_row = min(start_row + rows_per_batch - 1, max_row)
                try:
                    batch_values = await getRangeData(
                        client,
                        workbook_id,
                        sheet_id,
                        operator_id,
                        access_token,
                        start_row,
                        end_row,
                        max_column,
                    )
                    all_values.extend(batch_values)
                except Exception as e:
                    _raise_fetch(
                        f"分批获取数据失败（行 {start_row}-{end_row}）:\n"
                        f"{format_exception(e)}"
                    )
            return all_values

        try:
            return await getRangeData(
                client,
                workbook_id,
                sheet_id,
                operator_id,
                access_token,
                1,
                max_row,
                max_column,
            )
        except Exception as e:
            _raise_fetch(f"获取 Sheet 数据失败:\n{format_exception(e)}")


async def fetch_workbook_sheets(
    url: str,
    *,
    aegis_key: str | None = None,
    aegis_secret: str | None = None,
    workid: str | None = None,
) -> list[tuple[str, list[list[Any]]]]:
    workbook_id = extract_workbook_id_from_url(url)
    final_key = aegis_key or DEFAULT_AEGIS_KEY
    final_secret = aegis_secret or DEFAULT_AEGIS_SECRET
    final_workid = workid or DEFAULT_WORKID

    for attempt in range(2):
        try:
            access_token, operator_id = await getTokenAndOperatorId(
                final_key, final_secret, final_workid
            )
            break
        except httpx.HTTPStatusError as e:
            if attempt == 0 and is_invalid_auth_error(e):
                clear_token_cache(final_key, final_secret, final_workid)
                continue
            raise DingtalkExcelFetchError(str(e)) from e

    sheets = await getSheetList(workbook_id, operator_id, access_token)
    if not sheets:
        _raise_fetch("工作簿中没有 Sheet")

    result: list[tuple[str, list[list[Any]]]] = []
    for sheet in sheets:
        name = str(sheet.get("name") or "").strip()
        sheet_id = str(sheet.get("id") or "").strip()
        if not name or not sheet_id:
            continue
        values = await _fetch_sheet_values_raw(
            workbook_id=workbook_id,
            sheet_id=sheet_id,
            operator_id=operator_id,
            access_token=access_token,
        )
        result.append((name, values))
    return result


def _cell_nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def write_workbook_xlsx(
    sheets: list[tuple[str, list[list[Any]]]],
    output_path: Path,
) -> None:
    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        for row_idx, row in enumerate(rows, start=1):
            if not isinstance(row, list):
                continue
            for col_idx, value in enumerate(row, start=1):
                if _cell_nonempty(value):
                    ws.cell(row=row_idx, column=col_idx, value=value)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def guess_output_stem(sheets: list[tuple[str, list[list[Any]]]]) -> str:
    for _name, rows in sheets:
        for row in rows[:6]:
            if not isinstance(row, list):
                continue
            for cell in row:
                if cell is None:
                    continue
                text = str(cell)
                m = _VERSION_STEM_RE.search(text)
                if m:
                    return f"{m.group(1)}版本用例"
    return "钉钉版本用例"


def dingtalk_url_to_xlsx(
    url: str,
    output_path: Path,
    *,
    aegis_key: str | None = None,
    aegis_secret: str | None = None,
    workid: str | None = None,
) -> Path:
    sheets = asyncio.run(
        fetch_workbook_sheets(
            url,
            aegis_key=aegis_key,
            aegis_secret=aegis_secret,
            workid=workid,
        )
    )
    write_workbook_xlsx(sheets, output_path)
    return output_path
