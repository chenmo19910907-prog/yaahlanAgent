#!/usr/bin/env python3
"""从 2.4.9 家族相关 Excel 各 Sheet 生成 documents/家族改版.md（知识库）。"""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    return s


def find_header_row(rows: list[tuple]) -> int:
    for i, row in enumerate(rows):
        for c in row:
            if _cell_str(c) == "功能模块":
                return i
    return -1


def extract_meta(rows: list[tuple], header_idx: int) -> list[str]:
    lines: list[str] = []
    for r in rows[:header_idx]:
        parts = [_cell_str(x) for x in r[:6] if _cell_str(x)]
        if parts and any("设计人" in p or "测试工具" in p or "版本" in p for p in parts):
            line = " / ".join(parts[:4])
            if len(line) > 240:
                line = line[:237] + "…"
            lines.append("- " + line)
    return lines[:8]


def md_escape_line(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return s


def section_zh(idx_zero: int) -> str:
    m = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十"]
    n = idx_zero + 1
    if n <= 10:
        return m[n - 1]
    if n == 11:
        return "十一"
    if n == 12:
        return "十二"
    return str(n)


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("需要 openpyxl：python3 -m venv .venv && . .venv/bin/activate && pip install openpyxl", file=sys.stderr)
        return 1

    xlsx = Path(
        "/Users/user/Desktop/2.4.9版本用例（家族改版、活动大入口改为内嵌web、币商充值增加真人认证）.xlsx"
    )
    out = Path(__file__).resolve().parents[1] / "documents" / "家族改版.md"
    if len(sys.argv) >= 2:
        xlsx = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        out = Path(sys.argv[2])

    if not xlsx.is_file():
        print(f"找不到文件: {xlsx}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    sheet_order = list(wb.sheetnames)

    parts: list[str] = []
    parts.append("# 家族改版功能知识库（Yaahlan v2.4.9）\n\n")
    parts.append(f"> **文档来源**：`{xlsx}`  \n")
    parts.append(
        "> **整理方式**：按 Excel **各 Sheet** 阅读；表头行为「用例等级 / 功能模块 / case步骤描述 / 预期结果」；"
        "合并单元格导致的空「功能模块」按上行继承。正文仅保留步骤与预期，**不保留**各端 pass/fail 执行列。\n\n"
    )
    parts.append("## 目录\n\n")
    for i, sn in enumerate(sheet_order):
        idx = section_zh(i)
        short = sn.split("（")[0].strip()
        parts.append(f"- {idx}、{short}（Sheet：`{sn}`）\n")
    parts.append("\n---\n\n## 文档说明与维护\n\n")
    parts.append("| 项 | 说明 |\n|---|---|\n")
    parts.append("| 版本口径 | Yaahlan **v2.4.9** 家族改版相关用例（与 Excel 一致） |\n")
    parts.append("| 阅读建议 | 各章 `#####` 对应原表「功能模块」列，可折叠查阅。 |\n")
    parts.append("| 非家族说明 | Excel **文件名**含活动大入口、币商等版本主题；**当前工作簿内下列 Sheet 均为家族改版用例**（无其他业务 Sheet）。 |\n")
    parts.append("\n---\n")

    for si, sheet_name in enumerate(sheet_order):
        ws = wb[sheet_name]
        raw = list(ws.iter_rows(values_only=True))
        hi = find_header_row(raw)
        if hi < 0:
            parts.append(f"\n## （跳过）`{sheet_name}`：未识别表头\n")
            continue

        meta_lines = extract_meta(raw, hi)
        idx = section_zh(si)
        title_short = sheet_name.split("（")[0].strip()
        parts.append(f"\n## {idx}、{title_short}\n\n")
        parts.append(f"**对应 Excel Sheet**：`{sheet_name}`\n\n")
        if meta_lines:
            parts.append("#### 设计 / 工具摘录（表头上方）\n\n")
            for m in meta_lines:
                parts.append(m + "\n")
            parts.append("\n")

        parts.append("#### 用例要点（步骤 → 预期）\n\n")

        current_module = ""
        last_step_text: str | None = None
        pending_expects: list[str] = []

        def flush_step() -> None:
            nonlocal parts, last_step_text, pending_expects
            if last_step_text is None or not pending_expects:
                return
            parts.append(f"- **步骤**：{md_escape_line(last_step_text)}\n")
            for e in pending_expects:
                if e:
                    parts.append(f"  - **预期**：{md_escape_line(e)}\n")
            parts.append("\n")
            last_step_text = None
            pending_expects = []

        for row in raw[hi + 1 :]:
            if not row or all(_cell_str(c) == "" for c in row[:8]):
                continue
            level = _cell_str(row[0]) if len(row) > 0 else ""
            mod = _cell_str(row[1]) if len(row) > 1 else ""
            step = _cell_str(row[2]) if len(row) > 2 else ""
            expect = _cell_str(row[3]) if len(row) > 3 else ""

            if not mod and not step and not expect and not level:
                continue

            if mod:
                flush_step()
                current_module = mod
                safe = current_module.strip() or "（未命名模块）"
                parts.append(f"##### {safe}\n\n")
                last_step_text = None
                pending_expects = []

            if not current_module:
                continue

            if not expect and not step:
                continue

            if step:
                flush_step()
                last_step_text = step
                pending_expects = [expect] if expect else []
            elif expect:
                if last_step_text is None:
                    last_step_text = "（同模块补充步骤）"
                pending_expects.append(expect)

        flush_step()

    wb.close()

    text = "".join(parts)
    text = re.sub("\n{4,}", "\n\n\n", text)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(text, encoding="utf-8")
    print(f"Wrote {out} ({len(text)} chars)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
