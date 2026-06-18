#!/usr/bin/env python3
"""根据知识库分析结果，导出发版回归用例评审 Excel 到桌面。"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "缺少 openpyxl，请执行: pip install -r scripts/requirements-kb-sync.txt"
    ) from exc

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_MD = ROOT / "regression-kb" / "发版回归用例.md"
DEFAULT_XLSX_SOURCE = Path("/Users/user/Desktop/发版回归case.xlsx")
DEFAULT_OUTPUT = Path("/Users/user/Desktop/发版回归case_评审导出.xlsx")

STRATEGY_COL = "回归策略"
NOTE_COL = "调整说明"
ACTION_COL = "建议动作"
SOURCE_COL = "来源"

# 建议新增用例（分析结论）
NEW_CASES: list[dict[str, str | int | None]] = [
    {
        "级别": 1,
        "一级模块": "人脸认证",
        "二级模块": "主流程",
        "执行case": "真人认证提交/通过/拒绝/重试",
        "预期结果": "各状态展示与拦截逻辑正确",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "testcase-kb+回归零覆盖",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "人脸认证",
        "二级模块": "支付拦截",
        "执行case": "提现/大额支付触发实名认证拦截",
        "预期结果": "未认证用户被引导至认证流程",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "充值提现严重缺陷簇",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "主题房",
        "二级模块": "活动房",
        "执行case": "活动房列表→进房→活动条→退房主路径",
        "预期结果": "主路径可用，活动条与房间态正确",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "原回归仅弱覆盖",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "语音房",
        "二级模块": "跨房PK",
        "执行case": "跨房PK邀请→进行中→结束计分",
        "预期结果": "PK流程完整，分值与展示正确",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "原仅有乱斗PK",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "语音房",
        "二级模块": "基本功能",
        "执行case": "关房后 keep closed，不可再进入",
        "预期结果": "关闭状态保持，无法再次进房",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-4281 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "语音房",
        "二级模块": "婚礼房",
        "执行case": "婚礼房到点切换为婚礼房形态",
        "预期结果": "到点房间形态与婚礼能力正确",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-5914 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "语音房",
        "二级模块": "PK模式",
        "执行case": "10/12麦PK切换后麦位布局展示正确",
        "预期结果": "麦位人数与布局与模式一致",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-4029 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "语音房",
        "二级模块": "房间在线和成员列表",
        "执行case": "在线列表频繁刷新不崩溃",
        "预期结果": "列表刷新稳定无闪退",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-1523 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "支付",
        "二级模块": "谷歌充值",
        "执行case": "谷歌原生充值成功",
        "预期结果": "支付完成，钻石到账",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-8243 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "支付",
        "二级模块": "三方充值",
        "执行case": "三方充值成功/失败态",
        "预期结果": "结果页与到账/失败提示正确",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "bug-kb 充值严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "支付",
        "二级模块": "转账验证",
        "执行case": "钱包转账+上行短信验证",
        "预期结果": "验证通过后可转账成功",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "testcase-kb有规则、回归无",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "送礼，登录199或136号段执行",
        "二级模块": "礼物面板",
        "执行case": "打开礼物面板立即点选非默认礼物并送出",
        "预期结果": "送出的是所选礼物而非默认礼物",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-5236 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "送礼，登录199或136号段执行",
        "二级模块": "礼物面板",
        "执行case": "礼物面板切换tab后背包入口仍可见",
        "预期结果": "背包入口不消失，可正常切背包",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-6269 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "送礼，登录199或136号段执行",
        "二级模块": "群聊送礼",
        "执行case": "IM群聊（非家族）送礼成功",
        "预期结果": "收礼方收到礼物消息与展示",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "documents/gift 三场景-补群聊",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "送礼，登录199或136号段执行",
        "二级模块": "背包",
        "执行case": "聊天气泡购买后在背包展示",
        "预期结果": "背包可见已购道具",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-3599 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "消息帧",
        "二级模块": "账号与安全",
        "执行case": "换绑手机/邮箱输入错误验证码",
        "预期结果": "提示明确，非系统错误",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-8623 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "消息帧",
        "二级模块": "全部列表",
        "执行case": "消息列表下拉刷新好友在玩状态",
        "预期结果": "下拉后状态更新，无需切tab",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-8427 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "好友CP",
        "二级模块": "关系空间",
        "执行case": "切换账号后CP/好友空间数据隔离",
        "预期结果": "仅展示当前账号关系数据",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-3628 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "好友CP",
        "二级模块": "关系空间",
        "执行case": "CP/好友空间打开礼物面板并送礼",
        "预期结果": "礼物面板正常打开可送出",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-8082 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "首页-游戏帧",
        "二级模块": "jackaroo",
        "执行case": "jackaroo进房后座位展示与自动入座",
        "预期结果": "座位完整，可正常入座开局",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "YAAH-4285 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "首页-游戏帧",
        "二级模块": "概率游戏",
        "执行case": "概率游戏最小化/恢复不卡死右移",
        "预期结果": "最小化与恢复后布局正常可操作",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "bug-kb 游戏严重",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "首页-游戏帧",
        "二级模块": "进房链路",
        "执行case": "全服广播点击进入对应游戏/房间",
        "预期结果": "进入目标房间或游戏，非误开游戏页",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "bug-kb 遗留",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "登录注册",
        "二级模块": "H5/活动",
        "执行case": "悬浮窗或H5进入活动不丢失登录态",
        "预期结果": "活动页已登录，功能可用",
        STRATEGY_COL: "每版必测",
        NOTE_COL: "外网回测严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "安装",
        "二级模块": "网络",
        "执行case": "核心页弱网/断网提示与恢复",
        "预期结果": "有明确提示，网络恢复后可继续使用",
        STRATEGY_COL: "大版本/专项",
        NOTE_COL: "全库无覆盖",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "我的帧",
        "二级模块": "H5",
        "执行case": "MK WebView打开充值/活动/个人中心H5",
        "预期结果": "页面加载正常，业务可操作",
        STRATEGY_COL: "大版本/专项",
        NOTE_COL: "域名/H5类缺陷多",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "首页-房间帧",
        "二级模块": "活动",
        "执行case": "活动页进房/抽奖/奖励到账",
        "预期结果": "流程完成且奖励正确",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "bug-kb 榜单活动严重",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 1,
        "一级模块": "公会（在内网回测，使用内网账号16028887200）",
        "二级模块": "薪资",
        "执行case": "币商公会长赊账后币商钱包钻石一致",
        "预期结果": "赊账成功对应币商账户钻石增加",
        STRATEGY_COL: "触发式(内网)",
        NOTE_COL: "YAAH-5336 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "全服广播",
        "二级模块": "展示",
        "执行case": "阿语环境下大礼物全服广播语序RTL正确",
        "预期结果": "语序与RTL展示符合阿语习惯",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-8222 线上问题",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
    {
        "级别": 2,
        "一级模块": "登录注册",
        "二级模块": "注册",
        "执行case": "谷歌头像注册后个人资料页展示谷歌头像",
        "预期结果": "资料页展示注册所用头像",
        STRATEGY_COL: "隔版抽测",
        NOTE_COL: "YAAH-1068 严重缺陷",
        ACTION_COL: "新增",
        SOURCE_COL: "知识库分析",
    },
]


@dataclass
class RegressionRow:
    level: int | None
    l1: str
    l2: str
    step: str
    expected: str
    tested: int = 0
    pass_count: int = 0
    fail_count: int = 0
    fail_versions: str = ""
    strategy: str = ""
    note: str = ""
    action: str = "保留"
    source: str = "原表"


def parse_md_cases(md_path: Path) -> list[RegressionRow]:
    text = md_path.read_text(encoding="utf-8")
    rows: list[RegressionRow] = []
    current_l1 = ""
    current_l2 = ""
    buf: dict | None = None

    for line in text.splitlines():
        m = re.match(r"^### (.+?)（\d+）", line)
        if m:
            current_l1 = m.group(1).strip()
            current_l2 = ""
            continue
        m = re.match(r"^#### (.+)$", line)
        if m and not re.match(r"^#### P\d", line) and "历史失败" not in m.group(1):
            current_l2 = m.group(1).strip()
            continue
        m = re.match(r"^(?:##### |#### )(P\d) · (.+)$", line)
        if m:
            if buf:
                rows.append(_buf_to_row(buf, current_l1, current_l2))
            level_map = {"P0": 1, "P1": 2, "P2": 3}
            buf = {
                "level": level_map.get(m.group(1)),
                "step": m.group(2).strip(),
                "expected": "",
                "tested": 0,
                "pass_count": 0,
                "fail_count": 0,
                "fail_versions": "",
            }
            continue
        if buf and line.startswith("- **预期**："):
            buf["expected"] = line.split("：", 1)[-1].strip()
            continue
        if buf and "历史回归" in line:
            mt = re.search(r"共 (\d+) 版", line)
            mp = re.search(r"通过 (\d+)", line)
            mf = re.search(r"失败 (\d+)", line)
            if mt:
                buf["tested"] = int(mt.group(1))
            if mp:
                buf["pass_count"] = int(mp.group(1))
            if mf:
                buf["fail_count"] = int(mf.group(1))
        if buf and "曾失败版本" in line:
            buf["fail_versions"] = line.split("：", 1)[-1].strip()

    if buf:
        rows.append(_buf_to_row(buf, current_l1, current_l2))
    return rows


def _buf_to_row(buf: dict, l1: str, l2: str) -> RegressionRow:
    row = RegressionRow(
        level=buf.get("level"),
        l1=l1,
        l2=l2,
        step=buf["step"],
        expected=buf.get("expected", ""),
        tested=buf.get("tested", 0),
        pass_count=buf.get("pass_count", 0),
        fail_count=buf.get("fail_count", 0),
        fail_versions=buf.get("fail_versions", ""),
    )
    classify_row(row)
    return row


def classify_row(row: RegressionRow) -> None:
    """按知识库分析结论标注回归策略。"""
    l1, step, level = row.l1, row.step, row.level
    text = f"{l1} {row.l2} {step}"

    if row.fail_count > 0:
        row.strategy = "每版必测"
        row.note = f"历史FAIL {row.fail_count}次" + (
            f"（{row.fail_versions}）" if row.fail_versions else ""
        )
        row.action = "保留"
        return

    # 审核专项
    if l1 == "ios审核版本":
        row.strategy = "审核专项"
        row.note = "非每版常规回归，仅提审/审核包执行"
        row.action = "移出常规"
        return

    # 大版本
    if l1 == "语言切换":
        row.strategy = "大版本/专项"
        row.note = "多语言/RTL大改时执行"
        row.action = "移出常规"
        return
    if l1 == "安装" and "覆盖安装" in step:
        row.strategy = "大版本/专项"
        row.note = "包体/渠道变更时执行"
        row.action = "移出常规"
        return
    if l1 == "权限与push" and level != 1:
        row.strategy = "大版本/专项"
        row.note = "OS/SDK升级时执行"
        row.action = "移出常规"
        return

    # 删除/归档
    delete_keywords = (
        "定制火箭礼物，内网测试",
        "发送20麦位体验卡道具",
    )
    if any(k in step for k in delete_keywords):
        row.strategy = "删除/归档"
        row.note = "覆盖极少且内网/下发强依赖"
        row.action = "建议删除"
        return
    if l1 == "其他":
        row.strategy = "删除/归档"
        row.note = "兜底模块，建议拆分到具体域或删除"
        row.action = "建议删除"
        return
    if "道具礼物" == step and row.tested <= 3:
        row.strategy = "删除/归档"
        row.note = "低覆盖且与面板用例重复"
        row.action = "建议删除"
        return

    # 触发式
    trigger_l1 = ("首页弹窗", "公会（在内网回测，使用内网账号16028887200）")
    if l1 in trigger_l1:
        row.strategy = "触发式"
        row.note = "冷启动/内网账号/买量等条件触发时执行"
        row.action = "移出常规"
        return
    if l1 == "首页弹窗" or (
        l1 == "首页-交友帧（新）"
        and any(k in step for k in ("弹窗", "买量", "冷启动", "内网"))
    ):
        row.strategy = "触发式"
        row.note = "特殊账号或运营配置触发"
        row.action = "移出常规"
        return

    # 每版必测 - 核心模块 P0
    must_l1 = (
        "支付",
        "首页-游戏帧",
        "送礼，登录199或136号段执行",
        "登录注册",
        "动态帧",
        "消息帧",
    )
    if l1 in must_l1 and level == 1:
        row.strategy = "每版必测"
        row.note = "核心模块P0"
        row.action = "保留"
        return
    if l1 == "语音房" and level == 1:
        if any(
            k in text
            for k in ("进房", "上麦", "下麦", "开房", "退房", "送礼", "连击", "麦位", "发言")
        ):
            row.strategy = "每版必测"
            row.note = "语音房核心P0"
            row.action = "保留"
            return

    # 隔版抽测 - 稳定 P2 展示类
    if level == 3 and row.tested >= 30 and row.fail_count == 0:
        display_words = ("展示", "banner", "入口", "tab", "榜单", "刷新", "跳转", "点击")
        if l1 in ("我的帧", "首页-交友帧（新）", "vip等级页", "客服相关", "语音房") and any(
            w in step for w in display_words
        ):
            row.strategy = "隔版抽测"
            row.note = f"P2展示类，{row.tested}版全绿"
            row.action = "移出常规"
            return
    if l1 == "我的帧" and level == 3 and row.tested >= 40:
        if any(k in step for k in ("关于", "帮助", "反馈", "隐私", "协议", "勋章墙", "珍宝墙")):
            row.strategy = "隔版抽测"
            row.note = "我的帧静态/跳转P2"
            row.action = "移出常规"
            return
    if l1 == "语音房" and level == 3 and row.tested >= 45:
        if any(
            k in step
            for k in ("小时榜", "背景音乐", "家族入口", "切换社交", "房间头像", "密码房间")
        ):
            row.strategy = "隔版抽测"
            row.note = "语音房非核心设置P2"
            row.action = "移出常规"
            return
    if l1 == "vip等级页" and level == 3:
        row.strategy = "隔版抽测"
        row.note = "与VIP功能模块重复展示"
        row.action = "移出常规"
        return
    if l1 == "送礼，登录199或136号段执行" and "全房送礼" in step and row.tested < 20:
        row.strategy = "隔版抽测"
        row.note = "覆盖版本较少，可隔版"
        row.action = "移出常规"
        return

    # 默认：P0保留每版，P1/P2隔版或保留
    if level == 1:
        row.strategy = "每版必测"
        row.note = "P0默认每版"
        row.action = "保留"
    elif level == 2:
        row.strategy = "隔版抽测"
        row.note = "P1建议隔版或按改动执行"
        row.action = "移出常规"
    else:
        row.strategy = "隔版抽测"
        row.note = "P2建议隔版抽测"
        row.action = "移出常规"


def load_from_xlsx(xlsx_path: Path) -> list[RegressionRow] | None:
    if not xlsx_path.is_file():
        return None
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = "版本回归case" if "版本回归case" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(all_rows) < 3:
        return None

    rows: list[RegressionRow] = []
    current_l1 = ""
    current_l2 = ""
    for row in all_rows[2:]:
        if not row or len(row) < 5:
            continue
        level_raw, l1, l2, step, expected = row[0], row[1], row[2], row[3], row[4]
        if l1:
            current_l1 = str(l1).strip()
        if l2 and str(l2).strip() not in ("-", ""):
            current_l2 = str(l2).strip()
        if not step:
            continue
        step_text = str(step).strip()
        if not step_text:
            continue
        level = None
        if level_raw is not None and str(level_raw).strip().isdigit():
            level = int(str(level_raw).strip())
        r = RegressionRow(
            level=level,
            l1=current_l1,
            l2=current_l2,
            step=step_text,
            expected=str(expected).strip() if expected else "",
        )
        classify_row(r)
        rows.append(r)
    return rows


def write_workbook(
    existing_rows: list[RegressionRow],
    output_path: Path,
    *,
    source_label: str,
) -> None:
    wb = openpyxl.Workbook()

    # Sheet1: 用例评审
    ws = wb.active
    ws.title = "版本回归case"
    headers = [
        "级别",
        "一级模块",
        "二级模块",
        "执行case",
        "预期结果",
        STRATEGY_COL,
        NOTE_COL,
        ACTION_COL,
        SOURCE_COL,
        "历史执行版数",
        "历史通过",
        "历史失败",
        "曾失败版本",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in existing_rows:
        ws.append(
            [
                r.level,
                r.l1,
                r.l2 if r.l2 != "—" else "-",
                r.step,
                r.expected,
                r.strategy,
                r.note,
                r.action,
                r.source,
                r.tested or "",
                r.pass_count or "",
                r.fail_count or "",
                r.fail_versions,
            ]
        )

    for item in NEW_CASES:
        ws.append(
            [
                item.get("级别"),
                item.get("一级模块"),
                item.get("二级模块"),
                item.get("执行case"),
                item.get("预期结果"),
                item.get(STRATEGY_COL),
                item.get(NOTE_COL),
                item.get(ACTION_COL),
                item.get(SOURCE_COL),
                "",
                "",
                "",
                "",
            ]
        )

    # Sheet2: 统计汇总
    ws2 = wb.create_sheet("策略统计")
    ws2.append(["回归策略", "条数"])
    counter: dict[str, int] = {}
    action_counter: dict[str, int] = {}
    for r in existing_rows:
        counter[r.strategy] = counter.get(r.strategy, 0) + 1
        action_counter[r.action] = action_counter.get(r.action, 0) + 1
    for item in NEW_CASES:
        s = str(item.get(STRATEGY_COL, ""))
        counter[s] = counter.get(s, 0) + 1
        a = str(item.get(ACTION_COL, ""))
        action_counter[a] = action_counter.get(a, 0) + 1

    for k, v in sorted(counter.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["建议动作", "条数"])
    for k, v in sorted(action_counter.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.append([])
    ws2.append(["数据来源", source_label])
    ws2.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws2.append(["原有用例", len(existing_rows)])
    ws2.append(["建议新增", len(NEW_CASES)])

    # Sheet3: 说明
    ws3 = wb.create_sheet("使用说明")
    notes = [
        "本表由 auto-generate-testcase 根据 regression-kb / bug-kb 分析导出。",
        "",
        "回归策略列含义：",
        "  每版必测 — 每个发版版本常规双端回归应执行",
        "  隔版抽测 — 隔一个版本或按改动模块抽测",
        "  大版本/专项 — 包体/SDK/多语言等大改时执行",
        "  审核专项 — 仅 iOS 提审/审核包执行",
        "  触发式 — 特定账号/冷启动/内网条件触发时执行",
        "  删除/归档 — 建议从基线集移除或并入专项",
        "",
        "建议动作：保留 | 移出常规 | 建议删除 | 新增",
        "",
        "筛选建议：发版前筛选「每版必测」+ 本版改动模块 + 历史失败用例。",
    ]
    for n in notes:
        ws3.append([n])

    # column widths
    widths = [6, 28, 22, 48, 36, 14, 32, 12, 12, 12, 10, 10, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # color rows by action
    action_fills = {
        "建议删除": PatternFill("solid", fgColor="FCE4D6"),
        "移出常规": PatternFill("solid", fgColor="FFF2CC"),
        "新增": PatternFill("solid", fgColor="E2EFDA"),
    }
    for row_idx in range(2, ws.max_row + 1):
        action = ws.cell(row=row_idx, column=8).value
        fill = action_fills.get(str(action))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    ws.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--md", type=Path, default=DEFAULT_MD)
    parser.add_argument("--source-xlsx", type=Path, default=DEFAULT_XLSX_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    source_label = ""
    rows = load_from_xlsx(args.source_xlsx)
    if rows:
        source_label = str(args.source_xlsx)
        print(f"从 xlsx 加载 {len(rows)} 条: {args.source_xlsx}")
    else:
        if not args.md.is_file():
            print(f"错误: 找不到 {args.source_xlsx} 且找不到 {args.md}", file=__import__("sys").stderr)
            return 1
        rows = parse_md_cases(args.md)
        source_label = f"{args.md}（原 xlsx 不可用）"
        print(f"从 markdown 加载 {len(rows)} 条: {args.md}")

    write_workbook(rows, args.output, source_label=source_label)
    print(f"已导出: {args.output}")
    print(f"  原有用例: {len(rows)}，建议新增: {len(NEW_CASES)}，合计: {len(rows) + len(NEW_CASES)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
