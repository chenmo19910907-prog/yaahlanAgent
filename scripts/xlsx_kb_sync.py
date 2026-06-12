#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从版本用例表同步 testcase-kb（本地 xlsx 回退入口）。

**推荐**：`DingTalk/kb_sync_execute.py` 直接从钉钉目录拉取测试用例 Excel。

本脚本按版本顺序读取 xlsx，严格按：
1) 一次只处理一个文件
2) 文件内一次只处理一个 sheet
3) 当前 sheet 完成写入与优化后，才进入下一个 sheet
4) 当前文件全部 sheet 完成后，才进入下一个文件

冲突规则：同名「功能模块」章节以后处理版本覆盖（版本号更大为准）。
"""

from __future__ import annotations

import argparse
import sys
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

DEFAULT_SOURCE_DIR = Path("/Users/user/Desktop/未命名文件夹")
DEFAULT_OUTPUT_DOC_DIR = Path(__file__).resolve().parent.parent / "testcase-kb"

# 固定大模块文件映射（尽量少文件）
# 同步时跳过土语/俄语专项（与 kb_filter_locales.py 一致）
LOCALE_SKIP_RE = re.compile(
    r"土耳其政策整改|土语区分区策略",
    re.I,
)

KB_FILES: Dict[str, str] = {
    "room": "房间.md",
    "gift": "礼物.md",
    "family": "家族.md",
    "theme_room": "主题房.md",
    "moments": "动态.md",
    "message": "消息.md",
    "face_auth": "人脸认证.md",
    "auth_login": "注册登录.md",
    "customer_service": "客服.md",
    "super_admin": "超管.md",
    "agency": "公会.md",
    "coin": "币商.md",
    "game": "游戏.md",
    "rank_activity": "榜单与活动.md",
}


@dataclass
class CaseRow:
    module: str
    step: str
    expects: List[str]


def parse_version_from_filename(name: str) -> Optional[Tuple[int, int, int]]:
    m = re.search(r"(?:[vV])?\.?\s*(\d+)\s*\.\s*(\d+)\s*\.\s*(\d+)", name)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def normalize_text(s: str) -> str:
    return unicodedata.normalize("NFKC", s or "").strip()


def list_xlsx_files(source: Path) -> List[Tuple[Tuple[int, int, int], Path]]:
    items: List[Tuple[Tuple[int, int, int], Path]] = []
    for p in source.iterdir():
        if not p.is_file() or not p.name.lower().endswith(".xlsx"):
            continue
        if p.name.startswith("~$"):
            continue
        ver = parse_version_from_filename(p.name)
        if ver:
            items.append((ver, p))
    items.sort(key=lambda x: x[0])
    return items


def list_xlsx_paths(paths: List[Path]) -> List[Tuple[Tuple[int, int, int], Path]]:
    items: List[Tuple[Tuple[int, int, int], Path]] = []
    for raw in paths:
        p = raw.expanduser().resolve()
        if not p.is_file():
            raise SystemExit(f"文件不存在: {p}")
        if p.name.startswith("~$"):
            continue
        ver = parse_version_from_filename(p.name)
        if not ver:
            raise SystemExit(f"无法从文件名解析版本号: {p.name}")
        items.append((ver, p))
    items.sort(key=lambda x: x[0])
    return items


def detect_header_row_matrix(rows: List[List[Any]], max_scan: int = 80) -> Optional[int]:
    for i in range(min(len(rows), max_scan)):
        row = rows[i] if i < len(rows) else []
        vals = [normalize_text(str(row[c])) if c < len(row) and row[c] is not None else "" for c in range(12)]
        line = "".join(vals).lower()
        joined = "".join(vals)
        if "功能模块" in joined and ("步骤" in joined or "case" in line):
            return i
        # 早期版本表：案例编号 / 子功能名称 / 测试输入序列 / 预期结果
        if "子功能名称" in joined and "测试输入序列" in joined and "预期结果" in joined:
            return i
    return None


def find_column_indices(header: List[str]) -> Tuple[int, int, int, Optional[int]]:
    def pick(*keys: str) -> Optional[int]:
        for idx, h in enumerate(header):
            hh = h.replace(" ", "").lower()
            for k in keys:
                if k.lower().replace(" ", "") in hh:
                    return idx
        return None

    im = pick("功能模块", "子功能名称", "模块")
    is_ = pick("case步骤", "步骤描述", "测试步骤", "测试输入序列", "输入序列")
    ie = pick("预期结果", "期望结果")
    ig = pick("用例等级")

    return (1 if im is None else im, 2 if is_ is None else is_, 3 if ie is None else ie, ig)


def iter_cases_from_matrix(
    rows: List[List[Any]], sheet_title: str
) -> Tuple[str, List[CaseRow], Dict[str, str]]:
    hr = detect_header_row_matrix(rows)
    if hr is None:
        return sheet_title, [], {}

    from kb_version import extract_sheet_personnel

    sheet_personnel = extract_sheet_personnel(rows, hr)

    header_row = rows[hr] if hr < len(rows) else []
    max_col = max(len(header_row), 12)
    header = [normalize_text(str(header_row[c])) if c < len(header_row) and header_row[c] is not None else "" for c in range(max_col)]
    im, is_, ie, ig = find_column_indices(header)

    current_module = ""
    last_step = ""
    cases: List[CaseRow] = []

    for ridx in range(hr + 1, len(rows)):
        row = rows[ridx] if ridx < len(rows) else []

        def cell(ci: Optional[int]) -> str:
            if ci is None:
                return ""
            if ci >= len(row) or row[ci] is None:
                return ""
            return normalize_text(str(row[ci]).replace("\r", ""))

        mod_raw = cell(im)
        step = cell(is_)
        exp = cell(ie)
        grade = cell(ig)

        if not mod_raw and not step and not exp and not grade:
            continue
        if mod_raw and any(mod_raw.startswith(p) for p in ("设计人", "开发", "产品", "測試", "测试工具", "用例总计", "---")):
            continue

        if mod_raw:
            current_module = mod_raw

        if step and exp:
            last_step = step
            cases.append(CaseRow(module=current_module, step=step, expects=[exp]))
        elif step and not exp:
            last_step = step
            cases.append(CaseRow(module=current_module, step=step, expects=[]))
        elif (not step) and exp:
            if cases and cases[-1].module == current_module:
                cases[-1].expects.append(exp)
            else:
                cases.append(CaseRow(module=current_module, step=last_step or "（续）", expects=[exp]))

    return sheet_title, cases, sheet_personnel


def iter_sheet_matrices(path: Path) -> Iterable[Tuple[str, List[List[Any]]]]:
    """一次产出一个 sheet（严格逐 sheet 处理）。"""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            for name in wb.sheetnames:
                ws = wb[name]
                mr = int(ws.max_row or 1)
                mc = max(int(ws.max_column or 1), 12)
                rows: List[List[Any]] = []
                for r in ws.iter_rows(min_row=1, max_row=mr, max_column=mc, values_only=True):
                    rows.append(list(r))
                yield name, rows
        finally:
            wb.close()
        return
    except Exception:
        # openpyxl 读失败（如非法 sheet 名）回退 calamine
        from python_calamine import CalamineWorkbook

        book = CalamineWorkbook.from_path(str(path))
        for name in book.sheet_names:
            sh = book.get_sheet_by_name(name)
            yield name, sh.to_python(skip_empty_area=False)


def classify_big_module(file_name: str, sheet_title: str, module_name: str, step: str, exp: str) -> str:
    text = " ".join([file_name, sheet_title, module_name, step, exp]).lower()

    def has(*ks: str) -> bool:
        return any(k.lower() in text for k in ks)

    if has("主题房", "活动主题房"):
        return "theme_room"
    if has("家族", "公会家族"):
        return "family"
    if has("礼物", "gift", "背包", "盲盒", "勋章"):
        return "gift"
    if has("房间", "麦位", "进房", "语音房", "room"):
        return "room"
    if has("动态", "moment", "帖子", "热榜"):
        return "moments"
    if has("消息", "im", "私聊", "群聊", "聊天"):
        return "message"
    if has("人脸", "真人认证", "face", "认证"):
        return "face_auth"
    if has("登录", "注册", "注销", "账号与注册", "注册资料", "登录ui"):
        return "auth_login"
    if has("房间黑名单", "房间信息页") and has("黑名单"):
        return "room"
    if has("密码", "黑名单", "谁看过我", "设置页", "退出账号"):
        return "auth_login"
    if has("超管", "审核后台", "工单", "设备拉黑", "审核"):
        return "super_admin"
    if has("客服", "券包", "快捷回复"):
        return "customer_service"
    if has("am", "agency", "公会", "助理", "族长"):
        return "agency"
    if has("币商", "充值", "提现", "转账", "钻石", "钱包"):
        return "coin"
    if has("游戏", "pk", "大冒险", "battle"):
        return "game"
    if has("榜", "活动", "排名", "排行", "banner"):
        return "rank_activity"
    return "room"


def section_heading(module_name: str) -> str:
    return f"## 功能模块：{module_name}"


def ensure_doc(path: Path, title: str) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        f"# {title}\n\n"
        "- **说明**：按版本顺序从 xlsx 逐文件逐 sheet 汇总。\n"
        "- **冲突规则**：同名功能模块以后处理版本覆盖。\n"
        "\n---\n\n",
        encoding="utf-8",
    )


def replace_or_insert_module_section(content: str, module_name: str, body: str) -> str:
    heading = section_heading(module_name)
    new_block = f"{heading}\n\n{body.strip()}\n\n"
    pattern = re.compile(
        r"^## 功能模块：" + re.escape(module_name) + r"\s*$.*?^(?=## 功能模块：|\Z)",
        re.MULTILINE | re.DOTALL,
    )
    if pattern.search(content):
        return pattern.sub(lambda _m: new_block, content, count=1)
    return content.rstrip() + "\n\n" + new_block


def build_module_body(
    version_label: str,
    xlsx_name: str,
    sheet_title: str,
    module_name: str,
    rows: List[CaseRow],
    personnel: Optional[Dict[str, str]] = None,
) -> str:
    from kb_version import render_meta_header

    out: List[str] = []
    out.append(f"##### {sheet_title} · {module_name}")
    out.append("")
    out.append(render_meta_header(version_label, xlsx_name, personnel))
    out.append("")
    for c in rows:
        out.append(f"- **步骤**：{c.step}")
        if c.expects:
            for e in c.expects:
                out.append(f"  - **预期**：{e}")
        else:
            out.append("  - **预期**：_（表中未单列预期）_")
        out.append("")
    return "\n".join(out).strip()


def optimize_doc(path: Path) -> None:
    text = path.read_text(encoding="utf-8")
    text = text.replace("\r\n", "\n")
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    lines = [ln.rstrip() for ln in text.split("\n")]
    path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def process_one_sheet(
    source_name: str,
    version_label: str,
    sheet_name: str,
    rows: List[List[Any]],
) -> List[Path]:
    stitle, cases, sheet_personnel = iter_cases_from_matrix(rows, sheet_name)
    if not cases:
        return []

    grouped: Dict[Tuple[str, str], List[CaseRow]] = {}
    for c in cases:
        module_name = normalize_text(c.module) or "（未填功能模块）"
        big_key = classify_big_module(source_name, stitle, module_name, c.step, " ".join(c.expects))
        k = (big_key, module_name)
        grouped.setdefault(k, []).append(c)

    touched: List[Path] = []
    for (big_key, module_name), module_rows in grouped.items():
        out_name = KB_FILES[big_key]
        out_path = DEFAULT_OUTPUT_DOC_DIR / out_name
        ensure_doc(out_path, out_name.replace(".md", ""))

        body = build_module_body(
            version_label,
            source_name,
            stitle,
            module_name,
            module_rows,
            sheet_personnel,
        )
        old = out_path.read_text(encoding="utf-8")
        merged = replace_or_insert_module_section(old, module_name, body)
        out_path.write_text(merged, encoding="utf-8")
        touched.append(out_path)

    # 当前 sheet 完成后立刻优化（再进入下一个 sheet）
    for p in sorted(set(touched)):
        optimize_doc(p)

    return sorted(set(touched))


def process_workbook_sheets(
    source_name: str,
    version_label: str,
    sheets: Iterable[Tuple[str, List[List[Any]]]],
    *,
    skip_locale: bool = True,
) -> List[Path]:
    """处理一个工作簿的全部 sheet（钉钉 / 本地 xlsx 共用）。"""
    all_touched: List[Path] = []
    for sheet_name, matrix in sheets:
        if skip_locale and LOCALE_SKIP_RE.search(sheet_name):
            print(f"  -> 跳过 sheet（土语/俄语专项）: {sheet_name}")
            continue
        print(f"  -> sheet: {sheet_name}")
        touched = process_one_sheet(source_name, version_label, sheet_name, matrix)
        if touched:
            print(f"     已更新: {', '.join(p.name for p in touched)}")
            all_touched.extend(touched)
        else:
            print("     跳过（未识别到测试用例表头或无有效行）")
    return sorted(set(all_touched))


def reset_output_dir(out_dir: Path | None = None) -> None:
    target = out_dir or DEFAULT_OUTPUT_DOC_DIR
    target.mkdir(parents=True, exist_ok=True)
    for p in target.glob("*.md"):
        p.unlink()


def main() -> None:
    ap = argparse.ArgumentParser(description="逐文件逐sheet生成大模块知识库")
    ap.add_argument("--source", type=Path, default=DEFAULT_SOURCE_DIR)
    ap.add_argument(
        "--file",
        type=Path,
        action="append",
        default=[],
        help="指定版本用例 xlsx（可重复传入；按版本号升序处理）",
    )
    ap.add_argument("--only-version", type=str, default="")
    ap.add_argument("--reset", action="store_true", help="先清空 testcase-kb/*.md")
    args = ap.parse_args()

    if args.file:
        items = list_xlsx_paths(args.file)
    else:
        if not args.source.is_dir():
            raise SystemExit(f"源目录不存在: {args.source}（或用 --file 指定 xlsx）")
        items = list_xlsx_files(args.source)
    if args.only_version:
        t = tuple(int(x) for x in args.only_version.split("."))
        items = [(v, p) for v, p in items if v == t]

    if args.reset:
        reset_output_dir()

    print(f"将处理 {len(items)} 个文件，输出目录: {DEFAULT_OUTPUT_DOC_DIR}")

    for ver, xlsx in items:
        vlabel = f"v{ver[0]}.{ver[1]}.{ver[2]}"
        print(f"\n=== 文件开始: {xlsx.name} ({vlabel}) ===")

        # 一个文件内严格逐sheet
        sheet_idx = 0
        if LOCALE_SKIP_RE.search(xlsx.name):
            print(f"  跳过整文件（土语/俄语专项）: {xlsx.name}")
            continue

        process_workbook_sheets(xlsx.name, vlabel, iter_sheet_matrices(xlsx))

        print(f"=== 文件完成: {xlsx.name} ===")


if __name__ == "__main__":
    main()
