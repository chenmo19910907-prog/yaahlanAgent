#!/usr/bin/env python3
"""结合桌面发版回归 case 与知识库缺口，生成小版本/大版本二分回归用例 Excel。"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "缺少 openpyxl，请执行: .venv/bin/pip install -r scripts/requirements-kb-sync.txt"
    ) from exc

# 复用项目内分类与缺口用例
import sys

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))
from export_regression_case_review_xlsx import (  # noqa: E402
    NEW_CASES,
    RegressionRow,
    classify_row,
)

DEFAULT_SOURCE = Path.home() / "Desktop" / "工作簿1.xlsx"
DEFAULT_OUTPUT = Path.home() / "Desktop" / "发版回归case_优化版.xlsx"

DELETE_KEYWORDS = (
    "定制火箭礼物，内网测试",
    "发送20麦位体验卡道具",
)
# 小版本规范：我的帧仅保留核心 L1（钱包/资料/房间入口）
MY_FRAME_SMALL_KEEP = (
    "展示用户头像",
    "钱包外显",
    "已创建房间",
    "查看别人资料页",
    "点击进入商城",
)

VOICE_CORE_L1_KEYWORDS = (
    "直接上麦",
    "申请上麦",
    "主动下麦",
    "麦克风说话",
    "发送文字消息",
    "红包发送",
    "红包领取",
)

# 小版本：登录仅保留主路径（其余注册方式隔版抽测）
LOGIN_SMALL_KEEP = ("手机号登录", "手机号注册")

# 小版本：消息帧核心 IM（其余 L1 隔版）
MSG_SMALL_KEEP = (
    "全部、好友、任务",
    "1V1消息",
    "发送文字、表情、图片",
    "好友申请",
    "群成员列表",
    "发送文字、表情、图片、视频",
    "互相添加好友",
)

# 小版本：动态帧核心（列表+详情+发布）
MOMENT_SMALL_KEEP = (
    "滑动翻页",
    "点赞",
    "评论",
    "点击进入详情页",
    "可发布纯文字",
    "发布成功后跳转",
    "展示自己的全部动态",
    "支持滑动、点击切换",
    "只展示已关注和好友",
)

# 小版本：支付核心
PAY_SMALL_KEEP = (
    "me帧-钱包",
    "支持原生充值",
    "充值完成后",
)

# 小版本：送礼核心（L1 房间+私聊+家族+动态，排除内网/L4）
GIFT_SKIP = DELETE_KEYWORDS + ("道具礼物", "1v1送礼引导", "设置心愿礼物", "定制")

# 大版本增补：第三方登录抽样（不全跑 10 种）
LOGIN_LARGE_SAMPLE = (
    "FB注册",
    "google注册",
    "FB登录",
    "google登录",
    "appleID注册",
    "appleID登录",
)

# 大版本：语音房非核心 L1 关键词（设置/PK/神秘人/榜单）
VOICE_LARGE_L1_KEYWORDS = (
    "PK",
    "密码",
    "背景",
    "神秘人",
    "榜单",
    "小时榜",
    "分享房间",
    "编辑房间",
    "模式切换",
    "真心话",
)

VOICE_LARGE_L2_KEYWORDS = (
    "PK",
    "密码",
    "背景",
    "神秘人",
    "分享",
    "邀请",
    "踢",
    "禁言",
    "音乐",
    "家族",
    "榜单",
    "团战",
    "乱斗",
    "1V1",
    "切换社交",
    "头像",
)

# 大版本：我的帧 L2 主路径
MY_FRAME_LARGE_L2_KEYWORDS = (
    "编辑用户资料",
    "礼物墙",
    "勋章墙",
    "家族",
    "账单明细",
    "装扮",
    "邀请码",
    "账号安全",
)

# 大版本：VIP 仅 L1/L2 核心特权
VIP_LARGE_KEYWORDS = (
    "特权VIP",
    "签到",
    "谁看过我",
    "隐身",
    "VIP等级",
    "定制礼物",
    "VIP客服",
)

# 完全不纳入回归基线（小版本/大版本均不输出）
EXCLUDE_L1 = ("首页弹窗",)


@dataclass
class ParsedCase:
    level: int | None
    l1: str
    l2: str
    step: str
    expected: str
    tested: int = 0
    pass_count: int = 0
    fail_count: int = 0
    fail_versions: list[str] = field(default_factory=list)
    strategy: str = ""
    note: str = ""
    pool: str = ""  # 小版本 | 大版本 | 剔除
    source: str = "原表"


def mark_small(c: ParsedCase, note: str) -> None:
    c.pool = "小版本"
    c.strategy = "小版本"
    c.note = note


def mark_large(c: ParsedCase, note: str) -> None:
    c.pool = "大版本"
    c.strategy = "大版本"
    c.note = note


def mark_exclude(c: ParsedCase, note: str) -> None:
    c.pool = "剔除"
    c.strategy = "剔除"
    c.note = note


def norm_result(cell) -> str:
    if cell is None:
        return "skip"
    text = str(cell).strip()
    if not text or text == "-":
        return "skip"
    upper = text.upper()
    if "FAIL" in upper or "失败" in text:
        return "fail"
    if "PASS" in upper or "通过" in text:
        return "pass"
    return "other"


def load_cases_from_refined_sheets(xlsx_path: Path) -> list[ParsedCase]:
    """从优化版「全量对照」「建议归档」读取（无版本列历史）。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    cases: list[ParsedCase] = []
    for sheet_name in ("全量对照", "建议归档"):
        if sheet_name not in wb.sheetnames:
            continue
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5 or not row[3]:
                continue
            if row[0] and str(row[0]).startswith("小版本"):
                break
            level_raw, l1, l2, step, expected = row[0], row[1], row[2], row[3], row[4]
            level = None
            if level_raw is not None and str(level_raw).strip().isdigit():
                level = int(str(level_raw).strip())
            fail_count = 0
            fail_versions: list[str] = []
            if len(row) > 11 and row[11]:
                fv = str(row[11]).strip()
                if fv:
                    fail_versions = [v.strip() for v in fv.split(",") if v.strip()]
                    fail_count = len(fail_versions)
            if len(row) > 10 and row[10] and not fail_count:
                try:
                    fail_count = int(row[10])
                except (TypeError, ValueError):
                    pass
            tested = pass_count = 0
            if len(row) > 8 and row[8]:
                try:
                    tested = int(row[8])
                except (TypeError, ValueError):
                    pass
            if len(row) > 9 and row[9]:
                try:
                    pass_count = int(row[9])
                except (TypeError, ValueError):
                    pass
            source = str(row[7]).strip() if len(row) > 7 and row[7] else "原表"
            cases.append(
                ParsedCase(
                    level=level,
                    l1=str(l1).strip() if l1 else "",
                    l2=str(l2).strip() if l2 and str(l2) != "-" else "",
                    step=str(step).strip(),
                    expected=str(expected).strip() if expected else "",
                    tested=tested,
                    pass_count=pass_count,
                    fail_count=fail_count,
                    fail_versions=fail_versions,
                    source=source,
                )
            )
    wb.close()
    # 去重
    seen: set[tuple[str, str, str]] = set()
    unique: list[ParsedCase] = []
    for c in cases:
        k = case_key(c)
        if k in seen:
            continue
        seen.add(k)
        unique.append(c)
    return unique


def load_cases_with_history(xlsx_path: Path) -> list[ParsedCase]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = "版本回归case" if "版本回归case" in wb.sheetnames else wb.sheetnames[0]
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return []

    header = rows[0]
    version_cols: list[tuple[int, str]] = []
    for idx, name in enumerate(header):
        if idx < 5 or not name:
            continue
        version_cols.append((idx, str(name).strip()))

    cases: list[ParsedCase] = []
    current_l1 = ""
    current_l2 = ""

    for row in rows[2:]:
        if not row or len(row) < 5:
            continue
        level_raw, l1, l2, step, expected = row[0], row[1], row[2], row[3], row[4]
        if l1:
            current_l1 = str(l1).strip()
        if l2 and str(l2).strip() not in ("-", ""):
            current_l2 = str(l2).strip()
        if not step:
            continue
        step_text = str(step).strip()
        if not step_text:
            continue

        level: int | None = None
        if level_raw is not None and str(level_raw).strip().isdigit():
            level = int(str(level_raw).strip())

        tested = pass_count = fail_count = 0
        fail_versions: list[str] = []
        for col_i, ver in version_cols:
            if col_i >= len(row):
                continue
            status = norm_result(row[col_i])
            if status == "skip":
                continue
            tested += 1
            if status == "pass":
                pass_count += 1
            elif status == "fail":
                fail_count += 1
                fail_versions.append(ver)

        cases.append(
            ParsedCase(
                level=level,
                l1=current_l1,
                l2=current_l2,
                step=step_text,
                expected=str(expected).strip() if expected else "",
                tested=tested,
                pass_count=pass_count,
                fail_count=fail_count,
                fail_versions=fail_versions,
            )
        )
    return cases


def case_key(c: ParsedCase | dict) -> tuple[str, str, str]:
    if isinstance(c, ParsedCase):
        return (c.l1, c.l2, c.step)
    return (str(c["一级模块"]), str(c["二级模块"]), str(c["执行case"]))


def should_archive(c: ParsedCase) -> bool:
    if c.l1 == "其他":
        return True
    if any(k in c.step for k in DELETE_KEYWORDS):
        return True
    if c.step == "道具礼物":
        return True
    return False


def should_exclude(c: ParsedCase) -> bool:
    """不纳入小版本也不纳入大版本增补。"""
    if should_archive(c):
        return True
    if c.l1 in EXCLUDE_L1:
        return True
    # ios 审核专项纳入小版本，L3 也不剔除
    if c.l1 == "ios审核版本":
        return False
    if c.level == 3:
        return True
    step = c.step.strip()
    if not step or step == "-":
        return True
    if c.l1.startswith("公会") and c.level != 1:
        return True
    return False


def qualifies_for_large(c: ParsedCase) -> bool:
    """大版本增补准入：各模块主链路扩展，不含 L3 细项。"""
    text = f"{c.l1} {c.l2} {c.step}"

    if c.l1.startswith("VIP功能") or c.l1 == "vip等级页":
        return c.level in (1, 2) and any(k in text for k in VIP_LARGE_KEYWORDS)

    if c.l1 == "我的帧":
        return c.level == 2 and any(k in c.step for k in MY_FRAME_LARGE_L2_KEYWORDS)

    if c.l1 == "语音房":
        if c.level == 2:
            return any(k in text for k in VOICE_LARGE_L2_KEYWORDS)
        return c.level == 1 and any(k in text for k in VOICE_LARGE_L1_KEYWORDS)

    if c.l1 == "登录注册":
        return any(k in c.step for k in LOGIN_LARGE_SAMPLE)

    if c.l1 == "支付":
        return c.level == 2 or "币商" in text or "三方" in text or "checkout" in text.lower()

    if c.l1 == "送礼，登录199或136号段执行":
        if any(k in c.step for k in GIFT_SKIP):
            return False
        return c.level == 2

    if c.l1 == "消息帧":
        return c.level == 2

    if c.l1 == "动态帧":
        return c.level == 2 and any(
            k in c.step for k in ("发布", "编辑", "好友帖", "语音房", "发现列表前10")
        )

    if c.l1 == "首页-游戏帧":
        return c.level == 2 and any(
            k in c.step for k in ("ludo", "jackaroo", "台球", "概率", "水果", "匹配", "康乐")
        )

    if c.l1 == "首页-交友帧（新）":
        return c.level == 2 and "搜索" in c.step

    if c.l1.startswith("公会"):
        return c.level == 1 and any(k in c.step for k in ("薪资", "公会", "赊账", "成员"))

    if c.l1 == "首页-房间帧":
        return c.level == 2 and any(k in c.step for k in ("活动", "进房", "抽奖", "banner"))
    if c.l1 in ("好友CP", "全服广播"):
        return c.level == 2
    if c.l1 == "客服相关":
        return c.level == 2 and "在线" in c.step

    if c.l1 == "安装":
        return "覆盖安装" in c.step or "ipad" in c.step.lower()

    if c.l1 == "语言切换":
        return True
    if c.l1 in ("权限与push", "多账号切换", "财富魅力等级", "贵族等级"):
        return c.level in (1, 2)

    return False


def assign_pool(c: ParsedCase) -> None:
    """小版本 | 大版本增补 | 剔除（不输出）。"""
    text = f"{c.l1} {c.l2} {c.step}"

    if should_exclude(c):
        mark_exclude(c, "L3/内网细项/重复，不纳入基线")
        return

    # iOS 提审/审核包相关，纳入小版本
    if c.l1 == "ios审核版本":
        mark_small(c, "iOS 审核专项，小版本执行")
        return

    # 历史 FAIL → 小版本
    if c.fail_count > 0:
        mark_small(c, f"历史FAIL {c.fail_count}次（{', '.join(c.fail_versions)}）")
        return

    # ── 小版本判定（与原规则一致）──
    if c.l1 == "我的帧" and c.level == 1 and any(k in c.step for k in MY_FRAME_SMALL_KEEP):
        mark_small(c, "我的帧核心 L1")
        return

    if c.l1 == "语音房":
        if c.level == 1 and any(k in text for k in VOICE_CORE_L1_KEYWORDS):
            mark_small(c, "语音房核心 L1")
            return
        if c.level == 2 and "跨房PK" in c.step:
            mark_small(c, "跨房PK主链路")
            return

    if c.l1 == "登录注册" and any(k in c.step for k in LOGIN_SMALL_KEEP):
        mark_small(c, "登录注册主路径")
        return

    if c.l1 == "支付" and (
        any(k in c.step for k in PAY_SMALL_KEEP)
        or (c.level == 2 and "充值完成" in c.step)
    ):
        mark_small(c, "支付核心")
        return

    if c.l1 == "送礼，登录199或136号段执行":
        if c.level == 1 or (c.level == 2 and c.l2 in ("房间送礼", "1V1私信送礼")):
            if not any(k in c.step for k in GIFT_SKIP):
                mark_small(c, "送礼核心（房间/私聊/群聊）")
                return

    if c.l1 == "消息帧" and c.level == 1 and any(k in c.step for k in MSG_SMALL_KEEP):
        mark_small(c, "消息帧核心 IM")
        return

    if c.l1 == "动态帧" and any(k in c.step for k in MOMENT_SMALL_KEEP):
        mark_small(c, "动态帧核心")
        return

    if c.l1 == "首页-游戏帧" and (
        c.level == 1 or (c.level == 2 and "jackaroo" in c.step.lower())
    ):
        mark_small(c, "游戏帧核心")
        return

    if c.l1 == "首页-交友帧（新）" and c.level == 1:
        mark_small(c, "交友帧核心入口")
        return

    if c.l1 in ("安装", "主题房", "人脸认证", "好友CP", "全服广播"):
        if c.level == 1:
            mark_small(c, "核心模块 P0")
            return
        if c.level == 2 and c.l1 in ("好友CP", "人脸认证"):
            mark_small(c, "核心模块 P1")
            return

    if c.l1 == "安装" and c.level == 1:
        mark_small(c, "安装冒烟")
        return

    if c.level == 1 and c.l1 in (
        "客服相关",
        "首页-房间帧",
        "财富魅力等级",
        "贵族等级",
        "多账号切换",
        "权限与push",
    ):
        mark_small(c, "P0 默认小版本")
        return

    # ── 大版本增补准入 ──
    if qualifies_for_large(c):
        mark_large(c, "大版本模块主链路增补")
        return

    mark_exclude(c, "与小版本重复或覆盖价值低，不纳入基线")


def merge_knowledge_gaps(cases: list[ParsedCase]) -> list[ParsedCase]:
    existing = {case_key(c) for c in cases}
    added: list[ParsedCase] = []
    for item in NEW_CASES:
        key = (
            str(item["一级模块"]),
            str(item["二级模块"]),
            str(item["执行case"]),
        )
        if key in existing:
            continue
        level = item.get("级别")
        c = ParsedCase(
            level=int(level) if level else None,
            l1=key[0],
            l2=key[1],
            step=key[2],
            expected=str(item.get("预期结果", "")),
            source="知识库补充",
        )
        kb_note = str(item.get("调整说明", "bug-kb 缺口"))
        assign_pool(c)
        if c.pool != "剔除":
            c.note = f"{c.note}；{kb_note}" if kb_note else c.note
        added.append(c)
    return cases + added


def copy_sheet(src_wb, dst_wb, sheet_name: str) -> None:
    if sheet_name not in src_wb.sheetnames:
        return
    src = src_wb[sheet_name]
    dst = dst_wb.create_sheet(sheet_name)
    for row in src.iter_rows(values_only=True):
        dst.append(list(row) if row else [])


STRATEGY_FILLS = {
    "小版本": PatternFill("solid", fgColor="E2EFDA"),
    "大版本": PatternFill("solid", fgColor="DDEBF7"),
    "剔除": PatternFill("solid", fgColor="F4CCCC"),
}


def write_case_sheet(ws, cases: list[ParsedCase], *, title_note: str) -> None:
    headers = [
        "级别",
        "一级模块",
        "二级模块",
        "执行case",
        "预期结果",
        "版本类型",
        "调整说明",
        "来源",
        "历史执行版数",
        "历史通过",
        "历史失败",
        "曾失败版本",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in cases:
        ws.append(
            [
                c.level,
                c.l1,
                c.l2 if c.l2 else "-",
                c.step,
                c.expected,
                c.strategy,
                c.note,
                c.source,
                c.tested or "",
                c.pass_count or "",
                c.fail_count or "",
                ", ".join(c.fail_versions),
            ]
        )

    for row_idx in range(2, ws.max_row + 1):
        fill = STRATEGY_FILLS.get(str(ws.cell(row=row_idx, column=6).value or ""))
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    widths = [6, 30, 20, 48, 32, 12, 28, 10, 12, 10, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.cell(row=ws.max_row + 2, column=1, value=title_note)


def load_all_cases(source: Path) -> list[ParsedCase]:
    if not source.is_file():
        return []
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheetnames = wb.sheetnames
    wb.close()
    if "版本回归case" in sheetnames:
        cases = load_cases_with_history(source)
        if cases:
            return cases
    if "全量对照" in sheetnames:
        return load_cases_from_refined_sheets(source)
    # 简化五列表（如 工作簿1.xlsx / Sheet1）
    return load_cases_with_history(source)


def build_workbook(source: Path, output: Path) -> dict[str, int]:
    cases = load_all_cases(source)
    for c in cases:
        # 先用原分类补充 strategy（无历史时）
        row = RegressionRow(
            level=c.level,
            l1=c.l1,
            l2=c.l2,
            step=c.step,
            expected=c.expected,
            tested=c.tested,
            pass_count=c.pass_count,
            fail_count=c.fail_count,
            fail_versions=", ".join(c.fail_versions),
        )
        classify_row(row)
        assign_pool(c)
        # 有历史 FAIL 时 assign_pool 已覆盖 strategy
        if c.fail_count == 0 and c.source == "原表":
            pass  # keep assign_pool result

    # 仅当源表为 legacy 时补充知识库缺口（优化版全量对照已含补充项）
    if any(c.source == "原表" for c in cases):
        cases = merge_knowledge_gaps(cases)

    small = [c for c in cases if c.pool == "小版本"]
    large_only = [c for c in cases if c.pool == "大版本"]
    excluded = [c for c in cases if c.pool == "剔除"]
    original = [c for c in cases if c.source == "原表"]

    def sort_key(c: ParsedCase):
        return (c.level or 9, c.l1, c.l2, c.step)

    small.sort(key=sort_key)
    large_only.sort(key=sort_key)
    excluded.sort(key=sort_key)
    all_sorted = sorted([c for c in cases if c.pool != "剔除"], key=sort_key)

    wb_out = openpyxl.Workbook()
    ws1 = wb_out.active
    ws1.title = "小版本回归case"
    write_case_sheet(
        ws1,
        small,
        title_note=f"小版本 {len(small)} 条 · 小发版只跑本 Sheet",
    )

    ws2 = wb_out.create_sheet("大版本增补case")
    write_case_sheet(
        ws2,
        large_only,
        title_note=f"大版本增补 {len(large_only)} 条 · 大发版 = 小版本全部 + 本 Sheet 全部",
    )

    ws3 = wb_out.create_sheet("全量对照")
    write_case_sheet(
        ws3,
        all_sorted,
        title_note=f"全量 {len(all_sorted)} 条 · 含版本类型标注",
    )

    ws4 = wb_out.create_sheet("策略统计")
    ws4.append(["版本类型", "条数"])
    ws4.append(["小版本", len(small)])
    ws4.append(["大版本增补", len(large_only)])
    ws4.append(["大版本合计（小+增补）", len(small) + len(large_only)])
    ws4.append(["剔除（不执行）", len(excluded)])
    ws4.append([])
    ws4.append(["原表用例", len(original)])
    ws4.append(["知识库补充", len(cases) - len(original)])
    ws4.append(["源数据合计", len(cases)])
    ws4.append(["可执行合计", len(small) + len(large_only)])
    ws4.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    ws_ex = wb_out.create_sheet("剔除清单")
    write_case_sheet(
        ws_ex,
        excluded,
        title_note=f"剔除 {len(excluded)} 条 · L3/审核/内网细项，大小版本均不跑",
    )

    ws5 = wb_out.create_sheet("使用说明")
    for line in [
        f"发版回归用例 — 源表 {source.name} + bug-kb 缺口补充",
        "",
        "仅区分两种发版：",
        "",
        "【小版本】跑「小版本回归case」全部",
        "  · 原则上只回归客户端核心主链路",
        "  · 约 90～100 条，小优化 / bug 修复发版使用",
        "",
        "【大版本】跑「小版本回归case」+「大版本增补case」全部",
        "  · 新增玩法、大规模优化、全功能回归",
        "  · 增补已精简：去掉 L3 细项、内网重复项；iOS 审核专项在小版本",
        "  · 各模块仅保留主链路扩展（第三方登录抽样、VIP/语音房/我的帧核心扩展）",
        "",
        "【剔除清单】不纳入任何发版回归，仅供参考。",
        "",
        "排期与研发确认本版属小版本还是大版本，二选一执行对应 Sheet。",
    ]:
        ws5.append([line])

    # 复制原表 checklist 与规范（若存在）
    src_wb = openpyxl.load_workbook(source, data_only=True)
    for name in ("发版后checklist", "回归case执行规范"):
        copy_sheet(src_wb, wb_out, name)
    src_wb.close()

    output.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output)

    return {
        "small": len(small),
        "large": len(large_only),
        "excluded": len(excluded),
        "total": len(cases),
        "added": len(cases) - len(original),
    }


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(description="精炼发版回归用例 Excel")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    if not args.source.is_file():
        print(f"错误: 找不到源文件 {args.source}", file=sys.stderr)
        return 1

    stats = build_workbook(args.source, args.output)
    print(f"已生成: {args.output}")
    print(
        f"  小版本 {stats['small']} | 大版本增补 {stats['large']} | "
        f"大版本合计 {stats['small'] + stats['large']} | 剔除 {stats['excluded']}"
    )
    print(f"  原表 {stats['total'] - stats['added']} + 知识库补充 {stats['added']} = {stats['total']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
