#!/usr/bin/env python3
"""从「发版回归case.xlsx」同步版本回归用例到 regression-kb/（单文件）。"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "缺少 openpyxl，请执行: pip install -r scripts/requirements-kb-sync.txt"
    ) from exc

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path("/Users/user/Desktop/发版回归case.xlsx")
DEFAULT_OUTPUT = ROOT / "regression-kb"
OUTPUT_DOC = "发版回归用例.md"

# 同步时删除的旧版按模块拆分文件（若存在）
LEGACY_SPLIT_FILES = (
    "房间.md",
    "礼物.md",
    "家族.md",
    "主题房.md",
    "动态.md",
    "消息.md",
    "人脸认证.md",
    "注册登录.md",
    "客服.md",
    "超管.md",
    "公会.md",
    "充值提现转账.md",
    "游戏.md",
    "榜单与活动.md",
    "特权VIP.md",
    "贵族.md",
    "个人主页.md",
    "其他.md",
)


@dataclass
class RegressionCase:
    row_num: int
    level: int | None
    l1: str
    l2: str
    step: str
    expected: str
    tested: int = 0
    pass_count: int = 0
    fail_count: int = 0
    fail_versions: list[str] = field(default_factory=list)


def norm_result(cell) -> str:
    if cell is None:
        return "skip"
    text = str(cell).strip()
    if not text or text == "-":
        return "skip"
    upper = text.upper()
    if "FAIL" in upper or "失败" in text:
        return "fail"
    if "PASS" in upper or "通过" in text:
        return "pass"
    return "other"


def level_label(level: int | None) -> str:
    if level == 1:
        return "P0"
    if level == 2:
        return "P1"
    return "P2"


def anchor_slug(text: str) -> str:
    slug = "".join(ch if ch.isalnum() else "-" for ch in text)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug.strip("-") or "section"


def parse_cases(xlsx_path: Path) -> tuple[list[RegressionCase], list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = "版本回归case"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return [], []

    header = rows[0]
    version_cols: list[tuple[int, str]] = []
    for idx, name in enumerate(header):
        if idx < 5 or not name:
            continue
        version_cols.append((idx, str(name).strip()))

    cases: list[RegressionCase] = []
    current_l1 = ""
    current_l2 = ""

    for row_idx, row in enumerate(rows[2:], start=3):
        if not row or len(row) < 5:
            continue
        level_raw, l1, l2, step, expected = row[0], row[1], row[2], row[3], row[4]
        if l1:
            current_l1 = str(l1).strip()
        if l2 and str(l2).strip() not in ("-", ""):
            current_l2 = str(l2).strip()
        if not step:
            continue
        step_text = str(step).strip()
        if not step_text:
            continue

        level: int | None = None
        if level_raw is not None and str(level_raw).strip().isdigit():
            level = int(str(level_raw).strip())

        expected_text = str(expected).strip() if expected else ""

        tested = pass_count = fail_count = 0
        fail_versions: list[str] = []
        for col_i, ver in version_cols:
            if col_i >= len(row):
                continue
            status = norm_result(row[col_i])
            if status == "skip":
                continue
            tested += 1
            if status == "pass":
                pass_count += 1
            elif status == "fail":
                fail_count += 1
                fail_versions.append(ver)

        cases.append(
            RegressionCase(
                row_num=row_idx,
                level=level,
                l1=current_l1,
                l2=current_l2,
                step=step_text,
                expected=expected_text,
                tested=tested,
                pass_count=pass_count,
                fail_count=fail_count,
                fail_versions=fail_versions,
            )
        )

    version_names = [v for _, v in version_cols]
    return cases, version_names


def render_case_entry(case: RegressionCase, *, heading: str = "#####") -> list[str]:
    pri = level_label(case.level)
    lines = [
        f"{heading} {pri} · {case.step}",
        "",
        f"- **二级模块**：{case.l2 or '—'}",
        f"- **预期**：{case.expected or '—'}",
    ]
    if case.tested:
        summary = f"共 {case.tested} 版 · 通过 {case.pass_count}"
        if case.fail_count:
            summary += f" · **失败 {case.fail_count}**"
        lines.append(f"- **历史回归**：{summary}")
    if case.fail_versions:
        vers = "、".join(case.fail_versions[:6])
        if len(case.fail_versions) > 6:
            vers += f" 等 {len(case.fail_versions)} 个版本"
        lines.append(f"- **曾失败版本**：{vers}")
    lines.append("")
    return lines


def render_full_doc(
    cases: list[RegressionCase],
    source_path: Path,
    generated_at: str,
    version_names: list[str],
) -> str:
    p0 = sum(1 for c in cases if c.level == 1)
    p1 = sum(1 for c in cases if c.level == 2)
    has_fail = [c for c in cases if c.fail_count > 0]
    fail_ids = {id(c) for c in has_fail}

    # 保持 Excel 一级模块出现顺序
    l1_order: list[str] = []
    by_l1: dict[str, list[RegressionCase]] = defaultdict(list)
    for case in cases:
        key = case.l1 or "未分类"
        if key not in by_l1:
            l1_order.append(key)
        by_l1[key].append(case)

    lines = [
        "# 发版回归用例",
        "",
        "> **文档类型**：版本发版回归用例全集（由发版回归 Excel 同步，单文件不拆分）",
        f"> **数据来源**：`{source_path}`",
        f"> **生成时间**：{generated_at}",
        "> **用途**：发版/版本回归时查阅必测步骤、预期与历史执行情况",
        "",
        "## 概览",
        "",
        "| 指标 | 值 |",
        "|------|-----|",
        f"| 用例条数 | {len(cases)} |",
        f"| P0（级别1） | {p0} |",
        f"| P1（级别2） | {p1} |",
        f"| 历史曾失败 | {len(has_fail)} |",
        f"| 覆盖版本列数 | {len(version_names)} |",
        "",
        "## 目录",
        "",
    ]

    for l1_name in l1_order:
        anchor = anchor_slug(l1_name)
        lines.append(f"- [{l1_name}（{len(by_l1[l1_name])}）](#{anchor})")

    if has_fail:
        lines.extend(["", "- [历史失败用例（优先每版回归）](#历史失败用例优先每版回归)", ""])

    lines.extend(["", "## 历史失败用例（优先每版回归）", ""])
    if has_fail:
        for case in has_fail:
            lines.extend(render_case_entry(case, heading="####"))
    else:
        lines.append("（无）")
        lines.append("")

    lines.extend(["---", "", "## 用例正文", ""])

    for l1_name in l1_order:
        l1_cases = by_l1[l1_name]
        anchor = anchor_slug(l1_name)
        lines.append(f"### {l1_name}（{len(l1_cases)}）")
        lines.append("")

        # 二级模块保持出现顺序
        l2_order: list[str] = []
        by_l2: dict[str, list[RegressionCase]] = defaultdict(list)
        for case in l1_cases:
            l2_key = case.l2 or "—"
            if l2_key not in by_l2:
                l2_order.append(l2_key)
            by_l2[l2_key].append(case)

        for l2_name in l2_order:
            l2_cases = by_l2[l2_name]
            if l2_name != "—":
                lines.append(f"#### {l2_name}")
                lines.append("")
            for case in l2_cases:
                if id(case) in fail_ids:
                    continue
                lines.extend(render_case_entry(case))

    return "\n".join(lines).rstrip() + "\n"


def render_readme(
    cases: list[RegressionCase],
    source_path: Path,
    generated_at: str,
    version_names: list[str],
) -> str:
    p0_total = sum(1 for c in cases if c.level == 1)
    fail_total = sum(1 for c in cases if c.fail_count > 0)

    lines = [
        "# regression-kb · 发版回归用例知识库",
        "",
        "> **文档类型**：版本发版回归用例归档（**单文件**，不按业务模块拆分）",
        f"> **数据来源**：`{source_path}`",
        f"> **生成时间**：{generated_at}",
        "",
        "## 说明",
        "",
        "本目录由桌面 **发版回归case.xlsx** 自动同步。全部用例集中在：",
        "",
        f"- **[`{OUTPUT_DOC}`]({OUTPUT_DOC})** — 全部发版回归用例（与 Excel 一级/二级模块结构一致）",
        "",
        "用途：",
        "",
        "- 发版/版本回归时查阅必测用例与预期",
        "- 结合 [`bug-kb/`](../bug-kb/README.md)、[`online-kb/`](../online-kb/README.md) 对照历史问题",
        "- Agent 生成版本用例时参考已有回归覆盖面",
        "",
        "每条用例含：**P0/P1、一级/二级模块、步骤、预期、历史 PASS/FAIL**；曾失败用例在正文前单独列出。",
        "",
        "## 统计",
        "",
        "| 指标 | 值 |",
        "|------|-----|",
        f"| 用例总数 | {len(cases)} |",
        f"| P0（级别1） | {p0_total} |",
        f"| 历史曾失败 | {fail_total} |",
        f"| 覆盖版本列数 | {len(version_names)} |",
        "",
        "### 一级模块分布（文内章节）",
        "",
    ]
    for l1, count in Counter(c.l1 for c in cases).most_common():
        lines.append(f"- **{l1 or '未分类'}**：{count}")

    lines.extend(
        [
            "",
            "## 维护",
            "",
            "```bash",
            "python3 scripts/regression_kb_from_xlsx.py",
            "python3 scripts/regression_kb_from_xlsx.py --xlsx ~/Desktop/发版回归case.xlsx",
            "```",
            "",
        ]
    )
    return "\n".join(lines) + "\n"


def sync(
    xlsx_path: Path,
    output_dir: Path,
    *,
    dry_run: bool = False,
) -> int:
    if not xlsx_path.is_file():
        print(f"错误: 找不到 xlsx: {xlsx_path}", file=sys.stderr)
        return 1

    cases, version_names = parse_cases(xlsx_path)
    if not cases:
        print("错误: 未解析到用例", file=sys.stderr)
        return 1

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    readme = render_readme(cases, xlsx_path, generated_at, version_names)
    full_doc = render_full_doc(cases, xlsx_path, generated_at, version_names)

    if dry_run:
        print(f"[dry-run] 将写入 {output_dir}/{OUTPUT_DOC}（{len(cases)} 条用例）")
        return 0

    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "README.md").write_text(readme, encoding="utf-8")
    (output_dir / OUTPUT_DOC).write_text(full_doc, encoding="utf-8")

    removed = 0
    for name in LEGACY_SPLIT_FILES:
        legacy = output_dir / name
        if legacy.is_file():
            legacy.unlink()
            removed += 1

    print(
        f"已同步 {len(cases)} 条用例 → {output_dir}/{OUTPUT_DOC}"
        + (f"（已删除 {removed} 个旧拆分文件）" if removed else "")
    )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="发版回归 xlsx → regression-kb（单文件）")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="发版回归 case xlsx 路径")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出目录")
    parser.add_argument("--dry-run", action="store_true", help="仅预览不写文件")
    args = parser.parse_args()
    return sync(args.xlsx, args.output, dry_run=args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
