"""
客诉反馈数据分析脚本
用法: python3 analyze.py <xlsx路径>
输出: JSON 格式分析结果，供 Agent 读取后生成 Canvas 和写入钉钉文档
"""
import sys
import json
import openpyxl
from datetime import datetime
from collections import Counter, defaultdict


def is_invalid(r: dict) -> bool:
    res = r.get('result')
    if isinstance(res, str) and res.strip() in ('测试', '测试数据，忽略即可', '重复问题'):
        return True
    return False


def analyze(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            all_rows.append(dict(zip(headers, row)))

    invalid_rows = [r for r in all_rows if is_invalid(r)]
    rows = [r for r in all_rows if not is_invalid(r)]

    # 分类统计
    cat_count = Counter(r.get('category') for r in rows)
    bug_count = cat_count.get('is bug', 0) + cat_count.get('experience bug', 0) + cat_count.get('repeat bug', 0)
    not_bug_count = cat_count.get('not bug', 0)

    # 时间统计
    response_deltas, solved_deltas = [], []
    for r in rows:
        rd, resp, sol = r.get('report date'), r.get('response date'), r.get('solved date')
        if rd and resp and isinstance(rd, datetime) and isinstance(resp, datetime):
            d = (resp - rd).total_seconds() / 3600
            if d >= 0:
                response_deltas.append(d)
        if rd and sol and isinstance(rd, datetime) and isinstance(sol, datetime):
            d = (sol - rd).total_seconds() / 3600
            if d >= 0:
                solved_deltas.append(d)

    avg_resp = sum(response_deltas) / len(response_deltas) if response_deltas else 0
    avg_sol = sum(solved_deltas) / len(solved_deltas) if solved_deltas else 0

    # 超过24h
    over24 = []
    for r in rows:
        rd, sol = r.get('report date'), r.get('solved date')
        if rd and sol and isinstance(rd, datetime) and isinstance(sol, datetime):
            delta = (sol - rd).total_seconds() / 3600
            if delta > 24:
                over24.append({
                    'delta_h': round(delta, 1),
                    'report_date': str(rd)[:16],
                    'category': r.get('category', ''),
                    'desc': str(r.get('desc', ''))[:100],
                    'result': str(r.get('result', ''))[:150],
                })
    over24.sort(key=lambda x: -x['delta_h'])

    # 功能集中分析
    keywords = {
        '登录/账号': ['登录', '账号', '邮箱', '手机号', '绑定', '换绑', '找回', '密码'],
        '充值/支付': ['充值', '支付', '银行', 'usdt', 'USDT'],
        '真人认证': ['真人认证', '人脸', '认证', 'liveness'],
        '公会/家族': ['公会', '家族', '工会'],
        '风控': ['风控', '拦截', '封禁', '多开', '双开'],
        '礼物/特效': ['礼物', '特效', '动效'],
        '提现': ['提现', 'withdraw'],
        '网络/连接': ['网络', '卡顿', '加载', '连接'],
        '极速模式': ['极速模式'],
        'PK': ['PK榜', 'pk榜', 'PK奖'],
        '薪资/奖励': ['薪资', '奖励', '结算', '任务奖励'],
    }
    feature_counts = {}
    for func, kws in keywords.items():
        cnt = sum(1 for r in rows if any(k in str(r.get('desc', '')) + str(r.get('result', '')) for k in kws))
        feature_counts[func] = cnt
    feature_counts = dict(sorted(feature_counts.items(), key=lambda x: -x[1]))

    # Bug 明细
    bugs = []
    for r in rows:
        if r.get('category') in ('is bug', 'experience bug', 'repeat bug'):
            bugs.append({
                'type': r.get('category'),
                'desc': str(r.get('desc', ''))[:100],
                'result': str(r.get('result', ''))[:150],
            })

    return {
        'total': len(rows),
        'invalid_count': len(invalid_rows),
        'category_detail': dict(cat_count),
        'bug_count': bug_count,
        'not_bug_count': not_bug_count,
        'avg_response_h': round(avg_resp, 2),
        'avg_solve_h': round(avg_sol, 2),
        'over24': over24,
        'feature_counts': feature_counts,
        'bugs': bugs,
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('用法: python3 analyze.py <xlsx路径>')
        sys.exit(1)
    result = analyze(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))
